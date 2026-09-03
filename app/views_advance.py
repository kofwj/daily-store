"""垫资台账：门店填流水、管理员兑付/导出。"""

from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, timedelta
from io import BytesIO
from typing import Any, Dict, List

import openpyxl
from flask import flash, g, redirect, render_template, request, session, url_for
from openpyxl.styles import Alignment, Font, PatternFill

from . import db, db_core, sesame
from .helpers import (
    _xlsx_safe,
    accessible_stores,
    admin_required,
    company_names,
    login_required,
    pagination,
    parse_date,
    parse_int,
    pick_store,
    readonly_required,
    request_scope,
    viewer_only,
    xlsx_bytes,
    xlsx_response,
)
from .helpers import (
    month_end as _month_end,
)


def _form_from_row(row) -> Dict[str, str]:
    return {
        "advance_id": str(row["id"]),
        "biz_date": row["biz_date"],
        "phone": row["phone"] or "",
        "broadband": "" if not row["broadband"] else str(row["broadband"]),
        "rebate": "" if not row["rebate"] else str(row["rebate"]),
        "other": "" if not row["other"] else str(row["other"]),
        "note": row["note"] or "",
    }


def _empty_form(today: date) -> Dict[str, str]:
    return {
        "advance_id": "",
        "biz_date": today.isoformat(),
        "phone": "",
        "broadband": "",
        "rebate": "",
        "other": "",
        "note": "",
    }


def _sum_totals(maps):
    out = {"broadband": 0.0, "rebate": 0.0, "other": 0.0, "sesame": 0.0, "total": 0.0}
    for item in maps:
        for key in out:
            out[key] += float(item.get(key) or 0)
    return out


def _render_advance(conn, store, stores, form, is_admin, is_viewer, today_d, *, all_stores=False):
    month = today_d.replace(day=1)
    month_end = _month_end(month)
    scope = request_scope(stores)
    scoped_ids = None if (not all_stores or not scope["active"]) else scope["ids"]
    # 本月列表要全量：单店看本月全部（日期倒序，新的在前）；全店看全部再按门店/地市筛。
    # 不能用固定 200 上限，否则超 200 笔时列表和“未兑”数字都不对。
    if not all_stores:
        rows = db.list_advances(
            conn, store_id=store["id"], start=month, end=month_end,
            limit=100000, offset=0,
        )
        totals = db.advance_month_totals(conn, [store["id"]], today_d).get(
            store["id"], {"broadband": 0, "rebate": 0, "other": 0, "sesame": 0, "total": 0}
        )
    else:
        rows = db.list_all_advances(conn, month, month_end)
        if scoped_ids is not None:
            allow = set(scoped_ids)
            rows = [r for r in rows if int(r["store_id"]) in allow]
        kpi_ids = scope["ids"] if scope["active"] else [s["id"] for s in stores]
        totals = _sum_totals(db.advance_month_totals(conn, kpi_ids, today_d).values())
    totals["unpaid"] = sum(1 for r in rows if not int(r["paid"] or 0))
    return render_template(
        "advance.html",
        store=store,
        stores=stores,
        all_stores=all_stores,
        form=form,
        is_admin=is_admin,
        is_viewer=is_viewer,
        rows=rows,
        month=month,
        totals=totals,
        scope=scope,
    )


def register_advance(app) -> None:
    @app.route("/advance", methods=["GET", "POST"])
    @login_required
    def advance_page():
        with db.get_db() as conn:
            is_admin = g.user["role"] == "admin"
            raw_sid = (request.values.get("store_id") or "").strip()
            all_stores = is_admin and request.method == "GET" and raw_sid in ("", "all") and not request.args.get("advance_id")
            store, stores = pick_store(conn, None if all_stores else raw_sid)
            if store is None:
                flash("还没有可填的门店，请管理员先建店", "error")
                return render_template("empty.html")
            today_d = db.today_local()
            is_viewer = g.user["role"] in ("readonly", "city")
            form = _empty_form(today_d)
            if request.method == "GET":
                raw_id = request.args.get("advance_id") or ""
                aid = parse_int(raw_id, None)
                if aid:
                    row = db.get_advance(conn, aid, store["id"])
                    if row:
                        form = _form_from_row(row)
            if request.method == "POST":
                if is_viewer:
                    flash("只读账号不能填垫资，联系管理员。", "error")
                    return redirect(url_for("advance_page", store_id=store["id"]))
                raw_id = request.form.get("advance_id") or ""
                aid = parse_int(raw_id, None)
                biz_date = parse_date(request.form.get("biz_date"), today_d)
                phone = (request.form.get("phone") or "").strip()
                note = (request.form.get("note") or "").strip()
                form = {
                    "advance_id": raw_id,
                    "biz_date": biz_date.isoformat(),
                    "phone": phone,
                    "broadband": request.form.get("broadband") or "",
                    "rebate": request.form.get("rebate") or "",
                    "other": request.form.get("other") or "",
                    "note": note,
                }
                try:
                    broadband = db.parse_money(request.form.get("broadband"))
                    rebate = db.parse_money(request.form.get("rebate"))
                    other = db.parse_money(request.form.get("other"))
                except ValueError:
                    flash("金额请填数字，可留空。", "error")
                    return _render_advance(conn, store, stores, form, is_admin, is_viewer, today_d)
                same_month = biz_date.year == today_d.year and biz_date.month == today_d.month
                if not is_admin and biz_date > today_d:
                    flash("非管理员不能记未来日期。", "error")
                    return _render_advance(conn, store, stores, form, is_admin, is_viewer, today_d)
                if not is_admin and not same_month:
                    flash("门店只能记本月垫资。", "error")
                    return _render_advance(conn, store, stores, form, is_admin, is_viewer, today_d)
                if not is_admin and not phone:
                    flash("门店填写垫资必须带号码。", "error")
                    return _render_advance(conn, store, stores, form, is_admin, is_viewer, today_d)
                if not db.money_ok(broadband, rebate, other):
                    flash("三类金额至少填一项。", "error")
                    return _render_advance(conn, store, stores, form, is_admin, is_viewer, today_d)
                try:
                    db.record_advance(
                        conn,
                        store_id=store["id"],
                        user_id=g.user["id"],
                        biz_date=biz_date,
                        phone=phone,
                        broadband=broadband,
                        rebate=rebate,
                        other=other,
                        note=note,
                        advance_id=aid,
                    )
                except ValueError as exc:
                    if str(exc) == "paid_locked":
                        flash("已兑付的垫资不能改，先让管理员取消兑付。", "error")
                    elif str(exc) == "imported_locked":
                        flash("芝麻服务费是官方导入的，不能改。", "error")
                    else:
                        flash("这条垫资不存在。", "error")
                    return _render_advance(conn, store, stores, form, is_admin, is_viewer, today_d)
                flash("垫资已保存，可在下方本月记录里核对；填错了点「改」。", "ok")
                return redirect(url_for("advance_page", store_id=store["id"]))
            return _render_advance(conn, store, stores, form, is_admin, is_viewer, today_d, all_stores=all_stores)

    @app.route("/advance/delete", methods=["POST"])
    @login_required
    @viewer_only
    def advance_delete():
        store_id = request.form.get("store_id") or ""
        advance_id = request.form.get("advance_id") or ""
        try:
            sid = int(store_id)
            aid = int(advance_id)
        except (TypeError, ValueError):
            flash("参数不对。", "error")
            return redirect(url_for("advance_page"))
        with db.get_db() as conn:
            if not db.user_can_access_store(conn, g.user, sid):
                flash("没有这家店的权限。", "error")
                return redirect(url_for("advance_page"))
            row = db.get_advance(conn, aid, sid)
            if row is None:
                flash("这条垫资不存在，或已删除。", "error")
            elif (row["source"] or "") == "sesame" and g.user["role"] != "admin":
                flash("芝麻服务费是官方导入的，店员不能删。", "error")
            elif int(row["paid"] or 0) and not (g.user["role"] == "admin" and (row["source"] or "") == "sesame"):
                flash("已兑付的垫资不能删。", "error")
            else:
                allow_imported = g.user["role"] == "admin"
                try:
                    ok = db.delete_advance(
                        conn, aid, sid, user_id=g.user["id"], allow_imported=allow_imported
                    )
                except ValueError as exc:
                    if str(exc) == "imported_locked":
                        flash("芝麻服务费是官方导入的，店员不能删。", "error")
                        ok = False
                    else:
                        raise
                else:
                    flash("已删除这条垫资。" if ok else "这条垫资不存在，或已删除。", "ok" if ok else "error")
        return redirect(url_for("advance_page", store_id=sid))

    @app.route("/advance/pay")
    @admin_required
    def advance_pay():
        today_d = db.today_local()
        month = parse_date(request.args.get("month"), today_d.replace(day=1)).replace(day=1)
        month_end = min(_month_end(month), today_d) if month.year == today_d.year and month.month == today_d.month else _month_end(month)
        store_id = request.args.get("store_id", "")
        paid_raw = request.args.get("paid", "0")
        scope = request.args.get("scope", "today")
        if scope not in ("today", "month"):
            scope = "today"
        paid = None
        if paid_raw == "0":
            paid = 0
        elif paid_raw == "1":
            paid = 1
        start = today_d if scope == "today" else month
        end = today_d if scope == "today" else month_end
        with db.get_db() as conn:
            stores = accessible_stores(conn)
            city_scope = request_scope(stores)
            sid = int(store_id) if store_id.isdigit() else None
            if sid and not any(s["id"] == sid for s in stores):
                sid = None
            scoped_ids = None if (sid or not city_scope["active"]) else city_scope["ids"]
            inbox = db.advance_today_inbox(conn, today_d)
            month_inbox = db.advance_inbox(conn, month, month_end)
            if scoped_ids is not None:
                allow = set(scoped_ids)
                inbox = [r for r in inbox if int(r["store_id"]) in allow]
                month_inbox = [r for r in month_inbox if int(r["store_id"]) in allow]
            total = db.count_advances(
                conn, store_id=sid, start=start, end=end, paid=paid, store_ids=scoped_ids
            )
            page, pages = pagination(request.args.get("page"), total)
            rows = db.list_advances(
                conn,
                store_id=sid,
                start=start,
                end=end,
                paid=paid,
                limit=50,
                offset=(page - 1) * 50,
                store_ids=scoped_ids,
            )
            sums = db.advance_range_sums(
                conn, start=start, end=end, store_id=sid, store_ids=scoped_ids
            )
            return render_template(
                "advance_pay.html",
                stores=stores,
                rows=rows,
                month=month,
                today=today_d,
                store_id=str(sid or ""),
                paid=paid_raw,
                scope=scope,
                city_scope=city_scope,
                inbox=inbox,
                month_inbox=month_inbox,
                page=page,
                pages=pages,
                total=total,
                sums=sums,
            )

    @app.route("/advance/pay", methods=["POST"])
    @admin_required
    def advance_pay_post():
        action = request.form.get("action") or "pay"
        ids = request.form.getlist("advance_id")
        month = request.form.get("month") or ""
        store_id = request.form.get("store_id") or ""
        paid = request.form.get("paid") or "0"
        scope = request.form.get("scope") or "today"
        with db.get_db() as conn:
            n = db.set_advance_paid(
                conn,
                ids,
                paid=(action != "unpay"),
                user_id=g.user["id"],
            )
        if action == "unpay":
            flash(f"已取消兑付 {n} 笔。门店现在可以改这条。" if n else "没有可取消的记录。", "ok" if n else "error")
        else:
            flash(f"已兑付 {n} 笔。" if n else "没有可兑付的记录。", "ok" if n else "error")
        return redirect(url_for("advance_pay", month=month, store_id=store_id, paid=paid, scope=scope))

    def _sesame_week_range():
        today_d = db.today_local()
        default_start, default_end = sesame.week_span(today_d)
        start = parse_date(request.args.get("start"), default_start)
        end = parse_date(request.args.get("end"), sesame.week_span(start)[1])
        if end < start:
            start, end = end, start
        if (end - start).days > 31:
            end = start + timedelta(days=6)
        return start, end

    def _sesame_period():
        """周/月双模式取数窗口。月报锁定自然月（当月截止今天），周报周一到周日。"""
        today_d = db.today_local()
        mode = request.args.get("mode", "week")
        if mode not in ("week", "month"):
            mode = "week"
        if mode == "month":
            anchor = parse_date(request.args.get("start"), today_d).replace(day=1)
            last = sesame.month_span(anchor)[1]
            end = min(last, today_d)
            start = anchor
            prev_start = sesame.month_span(anchor - timedelta(days=1))[0]
            prev_end = sesame.month_span(prev_start)[1]
            next_start = last + timedelta(days=1)
            next_end = sesame.month_span(next_start)[1]
        else:
            start, end = _sesame_week_range()
            prev_start, prev_end = start - timedelta(days=7), end - timedelta(days=7)
            next_start, next_end = start + timedelta(days=7), end + timedelta(days=7)
        return start, end, prev_start, prev_end, next_start, next_end, mode

    @app.route("/advance/sesame/week")
    @readonly_required
    def sesame_week_page():
        start, end, prev_start, prev_end, next_start, next_end, mode = _sesame_period()
        with db.get_db() as conn:
            stores = accessible_stores(conn)
            cities = sorted({(s["city"] or "").strip() or "未分地市" for s in stores})
            areas = sorted({(s["area_manager"] or "").strip() for s in stores if (s["area_manager"] or "").strip()})
            city = (request.args.get("city") or "").strip()
            if city and city not in cities:
                city = ""
            area = (request.args.get("area") or "").strip()
            if area and area not in areas:
                area = ""
            scoped = [s for s in stores if ((s["city"] or "").strip() or "未分地市") == city] if city else stores
            if area:
                scoped = [s for s in scoped if (s["area_manager"] or "").strip() == area]
            scoped_ids = [int(s["id"]) for s in scoped]
            rows = sesame.sesame_week_rows(conn, scoped_ids, start, end)
            totals = sesame.sesame_week_totals(rows)
            # 通报表列：小天才 / AI手机 按类别，直降按原始档位；办理笔数按净数（扣费−退款）
            breakdown, tier_cols = sesame.sesame_tier_breakdown(conn, scoped_ids, start, end)
            for r in rows:
                b = breakdown.get(r["store_id"], {})
                r["cat_charges"] = b.get("cats", {})
                r["tier_charges"] = b.get("tiers", {})
            cat_totals = {
                "小天才直降": sum(int(r["cat_charges"].get("小天才直降") or 0) for r in rows),
                "AI手机": sum(int(r["cat_charges"].get("AI手机") or 0) for r in rows),
            }
            # 各类办理笔数已是净数（扣费−退款），合计直接相加
            tier_totals = {
                t: sum(int(r["tier_charges"].get(t) or 0) for r in rows) for t in tier_cols
            }
            copy_text = sesame.render_week_text(rows, totals, start, end, city, mode=mode)
            return render_template(
                "sesame_week.html",
                start=start,
                end=end,
                prev_start=prev_start,
                prev_end=prev_end,
                next_start=next_start,
                next_end=next_end,
                city=city,
                cities=cities,
                area=area,
                areas=areas,
                rows=rows,
                totals=totals,
                tier_cols=tier_cols,
                cat_totals=cat_totals,
                tier_totals=tier_totals,
                copy_text=copy_text,
                period_label=sesame.period_label(start, end, mode),
                mode=mode,
                is_admin=g.user["role"] == "admin",
            )

    @app.route("/advance/sesame/week.xlsx")
    @readonly_required
    def sesame_week_xlsx():
        start, end, _ps, _pe, _ns, _ne, mode = _sesame_period()
        with db.get_db() as conn:
            stores = accessible_stores(conn)
            city = (request.args.get("city") or "").strip()
            cities = sorted({(s["city"] or "").strip() or "未分地市" for s in stores})
            if city and city not in cities:
                city = ""
            area = (request.args.get("area") or "").strip()
            areas = sorted({(s["area_manager"] or "").strip() for s in stores if (s["area_manager"] or "").strip()})
            if area and area not in areas:
                area = ""
            scoped = [s for s in stores if ((s["city"] or "").strip() or "未分地市") == city] if city else stores
            if area:
                scoped = [s for s in scoped if (s["area_manager"] or "").strip() == area]
            scoped_ids = [int(s["id"]) for s in scoped]
            rows = sesame.sesame_week_rows(conn, scoped_ids, start, end)
            breakdown, tier_cols = sesame.sesame_tier_breakdown(conn, scoped_ids, start, end)
        header = ["门店", "地市", "净笔数", "扣费笔数", "扣费金额", "退款笔数", "退款金额", "净额"]
        data = [
            [r["name"], r["city"], r["n"], r["charge_n"], r["charge"], r["refund_n"], r["refund_abs"], r["net"]]
            for r in rows
        ]
        tag = "".join(ch for ch in city if ch.isalnum()) or "all"
        kind = "month" if mode == "month" else "week"
        filename = f"sesame_{kind}_{start.isoformat()}_{end.isoformat()}_{tag}.xlsx"
        book = xlsx_bytes(header, data, sheet="芝麻" + ("月报" if mode == "month" else "周报"))
        # 按档位 sheet：新用户芝麻直降的分档办理笔数（净数：扣费−退款）
        import openpyxl

        wb = openpyxl.load_workbook(BytesIO(book))
        tier_acc: dict = {}
        for r in rows:
            b = breakdown.get(r["store_id"], {})
            for t, n in b.get("tiers", {}).items():
                acc_t = tier_acc.setdefault(t, {"stores": 0, "n": 0, "net": 0.0})
                acc_t["stores"] += 1
                acc_t["n"] += n
                acc_t["net"] = round(acc_t["net"] + b.get("tiers_net", {}).get(t, 0.0), 2)
        ws = wb.create_sheet("按档位")
        ws.append(["档位", "门店数", "净办理笔数", "净额"])
        for t in tier_cols:
            acc_t = tier_acc.get(t, {"stores": 0, "n": 0, "net": 0.0})
            ws.append([f"{t}档", acc_t["stores"], acc_t["n"], acc_t["net"]])
        buf = BytesIO()
        wb.save(buf)
        return xlsx_response(buf.getvalue(), filename)

    @app.route("/advance/sesame", methods=["GET"])
    @admin_required
    def advance_sesame_page():
        preview = sesame.load_preview(session.get("sesame_token") or "")
        return render_template("advance_sesame.html", preview=preview)

    @app.route("/advance/sesame/preview", methods=["POST"])
    @admin_required
    def advance_sesame_preview():
        uploaded = request.files.get("sesame_file")
        if uploaded is None or not uploaded.filename:
            flash("请选择芝麻服务费明细 xlsx。", "error")
            return redirect(url_for("advance_sesame_page"))
        old = session.pop("sesame_token", None)
        if old:
            sesame.drop_preview(old)
        if not (uploaded.filename or "").lower().endswith(".xlsx"):
            flash("只支持 .xlsx 文件。", "error")
            return redirect(url_for("advance_sesame_page"))
        data = uploaded.read()
        try:
            rows = sesame.parse_sesame_xlsx(data)
        except ValueError as exc:
            flash(f"解析失败：{exc}", "error")
            return redirect(url_for("advance_sesame_page"))
        # 订单信息（可选）：用于按档位分类；不传则档位统计进「未分类」
        order_error = ""
        orders = []
        order_file = request.files.get("order_file")
        if order_file is not None and order_file.filename:
            if not (order_file.filename or "").lower().endswith(".xlsx"):
                order_error = "订单信息只支持 .xlsx，已忽略。"
            else:
                try:
                    with db.get_db() as conn:
                        rules = sesame.tier_rules(conn)
                    orders = sesame.parse_orders_xlsx(order_file.read())
                    for o in orders:
                        o["category"] = sesame.tier_category(o.get("order_title"), rules)
                except ValueError as exc:
                    order_error = f"订单信息解析失败：{exc}（已忽略，本次导入不分档位）"
        with db.get_db() as conn:
            stores = accessible_stores(conn)
            groups = sesame.classify_sesame_rows(conn, rows, stores)
        total_in = round(sum(r["amount"] for r in groups["ready"]), 2)
        # ready 必须存全量，否则确认时只导入截断的前 200 条，超出部分被静默丢弃。
        # 表格展示单独截前 200，计数/合计/按钮始终用全量。
        if orders:
            rules = sesame.tier_rules(conn)
            cat_by_order = {o["order_no"]: o["category"] for o in orders}
            tier_preview = Counter(
                cat_by_order.get(str(r["ext_id"]).removesuffix("_R"), "未分类")
                for r in groups["ready"]
            )
        else:
            tier_preview = {}
        preview = {
            "ready": groups["ready"],
            "ready_shown": groups["ready"][:200],
            "ready_total": total_in,
            "ready_count": len(groups["ready"]),
            "skipped": groups["skipped"][:50],
            "unmatched": groups["unmatched"][:50],
            "ignored": groups["ignored"][:50],
            "file_name": uploaded.filename,
            "row_count": len(rows),
            "orders": orders,
            "order_file_name": (order_file.filename or "") if order_file else "",
            "order_error": order_error,
            "tier_preview": dict(tier_preview),
        }
        session["sesame_token"] = sesame.save_preview(preview)
        if not groups["ready"]:
            flash("没有可导入的新流水（可能都已导入，或对不上门店）。", "error")
        return redirect(url_for("advance_sesame_page"))

    @app.route("/advance/sesame/confirm", methods=["POST"])
    @admin_required
    def advance_sesame_confirm():
        token = session.pop("sesame_token", None)
        preview = sesame.load_preview(token or "")
        sesame.drop_preview(token or "")
        if not preview or not preview.get("ready"):
            flash("预览已过期，请重新上传。", "error")
            return redirect(url_for("advance_sesame_page"))
        ready = preview["ready"]
        orders = preview.get("orders") or []
        with db.get_db() as conn:
            db_core.begin_immediate(conn)  # 整批导入=一个长写事务，抢先占写锁防 DEFERRED 升级锁冲突
            # 再次校验未导入，避免并发重复
            stores = accessible_stores(conn)
            stores_by_id = {int(s["id"]): s for s in stores}
            rows = [{**r, "store_id": int(r["store_id"])} for r in ready if int(r.get("store_id") or 0) in stores_by_id]
            groups = sesame.classify_sesame_rows(conn, rows, stores)
            n = sesame.import_sesame_rows(conn, groups["ready"], user_id=g.user["id"])
            if orders:
                db_core.sesame_orders_upsert(conn, orders, db_core._now())
        extra = f"订单信息 {len(orders)} 条已更新档位分类。" if orders else ""
        flash((f"已导入 {n} 笔芝麻服务费。" if n else "没有导入新流水。") + extra, "ok" if n else "error")
        return redirect(url_for("advance_page"))

    @app.route("/advance.xlsx")
    @admin_required
    def advance_xlsx():
        today_d = db.today_local()
        month = parse_date(request.args.get("month"), today_d.replace(day=1)).replace(day=1)
        month_end = _month_end(month)
        with db.get_db() as conn:
            stores = accessible_stores(conn)
            rows = db.list_all_advances(conn, month, month_end)
            data = _build_advance_xlsx(stores, rows, month, conn)
        return xlsx_response(data, f"advance_{month.strftime('%Y_%m')}.xlsx")


def _build_advance_xlsx(stores, rows, month: date, conn=None) -> bytes:
    organ_names = company_names(conn)
    by_store: Dict[int, List[Any]] = defaultdict(list)
    for row in rows:
        by_store[int(row["store_id"])].append(row)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill("solid", fgColor="E2EFDA")
    center = Alignment(horizontal="center", vertical="center")
    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "汇总表"
    summary["A1"] = _xlsx_safe(f"{month.month}月门店运营中心移动垫资费用报销汇总")
    summary["A1"].font = Font(bold=True, size=14)
    summary.merge_cells("A1:F1")
    for col, text in enumerate(["门店", "宽带调测费", "购机让利", "其他业务垫资", "芝麻服务费", "合计"], start=1):
        cell = summary.cell(2, col, text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    nt_rows: List[int] = []
    tz_rows: List[int] = []
    excel_row = 3
    for store in stores:
        sid = int(store["id"])
        items = by_store.get(sid, [])
        broadband = round(sum(float(r["broadband"] or 0) for r in items), 2)
        rebate = round(sum(float(r["rebate"] or 0) for r in items), 2)
        other = round(sum(float(r["other"] or 0) for r in items), 2)
        sesame = round(sum(float(r["sesame"] or 0) for r in items), 2)
        summary.cell(excel_row, 1, _xlsx_safe(store["name"]))
        summary.cell(excel_row, 2, broadband)
        summary.cell(excel_row, 3, rebate)
        summary.cell(excel_row, 4, other)
        summary.cell(excel_row, 5, sesame)
        summary.cell(excel_row, 6, f"=SUM(B{excel_row}:E{excel_row})")
        city = (store["city"] or "")
        if "泰州" in city:
            tz_rows.append(excel_row)
        else:
            nt_rows.append(excel_row)
        excel_row += 1
        _write_store_sheet(wb, store, items, header_fill, header_font, total_fill)
    if nt_rows:
        summary.cell(excel_row, 1, _xlsx_safe(organ_names["nt"]))
        for col, letter in enumerate(["B", "C", "D", "E", "F"], start=2):
            joined = "+".join(f"{letter}{r}" for r in nt_rows)
            cell = summary.cell(excel_row, col, "=" + joined)
            cell.fill = total_fill
            cell.font = Font(bold=True)
        excel_row += 1
    if tz_rows:
        summary.cell(excel_row, 1, _xlsx_safe(organ_names["tz"]))
        for col, letter in enumerate(["B", "C", "D", "E", "F"], start=2):
            joined = "+".join(f"{letter}{r}" for r in tz_rows)
            cell = summary.cell(excel_row, col, "=" + joined)
            cell.fill = total_fill
            cell.font = Font(bold=True)
        excel_row += 1
    if nt_rows or tz_rows:
        last = excel_row - 1
        first_company = last - (1 if nt_rows and tz_rows else 0)
        summary.cell(excel_row, 1, _xlsx_safe("门店运营中心"))
        for col, letter in enumerate(["B", "C", "D", "E", "F"], start=2):
            cell = summary.cell(excel_row, col, f"={letter}{first_company}+{letter}{last}" if nt_rows and tz_rows else f"={letter}{last}")
            cell.fill = total_fill
            cell.font = Font(bold=True)
    for col, width in enumerate([40, 14, 12, 14, 12, 12], start=1):
        summary.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_store_sheet(wb, store, items, header_fill, header_font, total_fill) -> None:
    raw = (store["short_name"] or store["name"] or "门店").strip()
    name = raw[:31] or "门店"
    base, n = name, 2
    while name in wb.sheetnames:
        name = f"{base[:28]}_{n}"
        n += 1
    ws = wb.create_sheet(name)
    ws["A1"] = _xlsx_safe(store["name"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:J1")
    headers = ["日期", "号码", "宽带调测费", "购机让利", "其他业务", "芝麻服务费", "合计", "备注", "是否报销", "报销日期"]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(2, col, text)
        cell.fill = header_fill
        cell.font = header_font
    start = 3
    for i, row in enumerate(items):
        r = start + i
        ws.cell(r, 1, _xlsx_safe(row["biz_date"]))
        ws.cell(r, 2, _xlsx_safe(row["phone"] or ""))
        ws.cell(r, 3, float(row["broadband"] or 0) or None)
        ws.cell(r, 4, float(row["rebate"] or 0) or None)
        ws.cell(r, 5, float(row["other"] or 0) or None)
        ws.cell(r, 6, float(row["sesame"] or 0) or None)
        ws.cell(r, 7, f"=SUM(C{r}:F{r})")
        ws.cell(r, 8, _xlsx_safe(row["note"] or ""))
        ws.cell(r, 9, _xlsx_safe("是" if int(row["paid"] or 0) else ""))
        ws.cell(r, 10, _xlsx_safe(row["paid_at"] or ""))
    end = start + max(len(items), 1) - 1
    total_row = end + 1
    ws.cell(total_row, 1, "合计")
    for col, letter in enumerate(["C", "D", "E", "F", "G"], start=3):
        cell = ws.cell(total_row, col, f"=SUM({letter}{start}:{letter}{end})")
        cell.fill = total_fill
        cell.font = Font(bold=True)
    widths = [12, 14, 12, 12, 12, 12, 10, 36, 10, 12]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
