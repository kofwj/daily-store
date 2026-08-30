"""芝麻服务费官方明细：解析、对店、导入垫资。"""

from __future__ import annotations

import json
import re
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from secrets import token_hex
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence

from openpyxl import load_workbook

from . import db_advances, db_core

# 档位分类：可从设置页改（app_meta: sesame_tier_rules），缺省 500=小天才、1000=AI手机
TIER_UNMATCHED = "未分类"
TIER_DEFAULT_RULES = {"xtc": 500, "ai": 1000}


def tier_rules(conn) -> Dict[str, int]:
    raw = db_core.get_setting(conn, "sesame_tier_rules", "")
    rules = dict(TIER_DEFAULT_RULES)
    if raw:
        try:
            data = json.loads(raw)
            for key in ("xtc", "ai"):
                v = int(data.get(key, rules[key]))
                if v > 0:
                    rules[key] = v
        except (TypeError, ValueError, json.JSONDecodeError):
            pass
    if rules["xtc"] == rules["ai"]:
        rules = dict(TIER_DEFAULT_RULES)
    return rules


def tier_category(title: str, rules: Mapping[str, int]) -> str:
    """从订单标题提取档位并归类：500=小天才、1000=AI手机、其余=新用户芝麻直降。"""
    m = re.search(r"(\d+)\s*档", str(title or ""))
    if not m:
        return TIER_UNMATCHED
    n = int(m.group(1))
    if n == int(rules["xtc"]):
        return "小天才直降"
    if n == int(rules["ai"]):
        return "AI手机"
    return "新用户芝麻直降"


def parse_orders_xlsx(data: bytes) -> List[Dict[str, Any]]:
    """解析芝麻订单信息 xlsx：订单号 + 档位类别，不碰姓名/手机号/身份证等隐私列。"""
    if not data:
        raise ValueError("没有文件")
    if len(data) > MAX_IMPORT_BYTES:
        raise ValueError("文件不能超过 4 MiB")
    try:
        wb = load_workbook(BytesIO(data), data_only=True, read_only=False)
    except Exception as exc:  # noqa: BLE001
        raise ValueError("打不开这份订单信息 Excel") from exc
    try:
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, max_row=max(ws.max_row or 1, 20), max_col=25, values_only=True))
    finally:
        wb.close()
    header_idx = None
    for i, r in enumerate(rows[:10]):
        labels = [str(v or "").strip() for v in r]
        if "订单号" in labels and "订单标题" in labels:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("不是芝麻订单信息（缺订单号 / 订单标题）")
    header = [str(v or "").strip() for v in rows[header_idx]]
    out: List[Dict[str, Any]] = []
    seen = set()
    for r in rows[header_idx + 1 :]:
        rec = dict(zip(header, r))
        order_no = str(rec.get("订单号") or "").strip()
        if not order_no or order_no in seen:
            continue
        seen.add(order_no)
        title = str(rec.get("订单标题") or "").strip()
        frozen_raw = rec.get("冻结金额")
        try:
            frozen = round(float(frozen_raw or 0), 2)
        except (TypeError, ValueError):
            frozen = 0.0
        try:
            terms = int(rec.get("期数") or 0)
        except (TypeError, ValueError):
            terms = 0
        out.append(
            {
                "order_no": order_no[:40],
                "store_code": str(rec.get("门店编码") or "").strip()[:40],
                "frozen": frozen,
                "terms": terms,
                "order_title": title[:120],
                "tier": (re.search(r"(\d+)\s*档", title).group(1) if re.search(r"(\d+)\s*档", title) else ""),
            }
        )
    if not out:
        raise ValueError("订单信息里没有数据行")
    return out

MAX_IMPORT_BYTES = 4 * 1024 * 1024


def mobile_code_of(raw: Any) -> str:
    text = str(raw or "").strip().upper()
    if text.startswith("JSCM_"):
        text = text[5:]
    return text.strip()


def parse_created(raw: Any) -> Optional[date]:
    if raw is None or raw == "":
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    text = str(raw).strip()
    if text.endswith(".0"):
        text = text[:-2]
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _cell(row: Sequence[Any], idx: int) -> Any:
    return row[idx] if idx < len(row) else None


def parse_sesame_xlsx(data: bytes) -> List[Dict[str, Any]]:
    if not data:
        raise ValueError("没有文件")
    if len(data) > MAX_IMPORT_BYTES:
        raise ValueError("文件不能超过 4 MiB")
    try:
        # 官方表标题行合并 A1:Q1，read_only 会把维度看成 1x1，只读到标题。
        wb = load_workbook(BytesIO(data), data_only=True, read_only=False)
    except Exception as exc:
        raise ValueError("打不开这份 Excel") from exc
    try:
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, max_row=max(ws.max_row or 1, 20), max_col=20, values_only=True))
    finally:
        wb.close()
    header_idx = None
    for i, row in enumerate(rows[:20]):
        labels = [str(v or "").strip() for v in row]
        if "流水号" in labels and "门店编码" in labels and "服务费金额" in labels:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("不是芝麻服务费明细（缺流水号 / 门店编码 / 服务费金额）")
    out: List[Dict[str, Any]] = []
    seen = set()
    for row in rows[header_idx + 1 :]:
        ext_id = str(_cell(row, 0) or "").strip()
        if not ext_id:
            continue
        if ext_id in seen:
            continue
        seen.add(ext_id)
        fee_raw = _cell(row, 13)
        try:
            platform_fee = db_advances.parse_money(fee_raw)
        except ValueError as exc:
            raise ValueError(f"流水 {ext_id} 的服务费不是数字") from exc
        try:
            order_amt = db_advances.parse_money(_cell(row, 12))
        except ValueError:
            order_amt = 0.0
        biz_date = parse_created(_cell(row, 16))
        note = str(_cell(row, 15) or "")
        refund = "退款" in note or ext_id.endswith("_R")
        # 方向由退款标志定，不赌表内符号
        sign = -1.0 if refund else  1.0  # 退款退回→垫资为负；正常→垫资为正
        out.append(
            {
                "ext_id": ext_id[:80],
                "order_no": str(_cell(row, 1) or "").strip()[:80],
                "city": str(_cell(row, 3) or "").strip(),
                "area": str(_cell(row, 4) or "").strip(),
                "merchant": str(_cell(row, 7) or "").strip(),
                "store_code": str(_cell(row, 8) or "").strip(),
                "mobile_code": mobile_code_of(_cell(row, 8)),
                "platform_store": str(_cell(row, 9) or "").strip(),
                "month": str(_cell(row, 11) or "").strip(),
                "order_amt": order_amt,
                "platform_fee": platform_fee,
                "amount": round(sign * abs(platform_fee), 2),
                "status": str(_cell(row, 14) or "").strip(),
                "note": note,
                "refund": refund,
                "biz_date": biz_date.isoformat() if biz_date else "",
            }
        )
    if not out:
        raise ValueError("明细里没有流水")
    return out


def _store_map(stores: Iterable[Any]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for store in stores:
        code = (store["mobile_code"] or "").strip()
        if code:
            out[code] = store
    return out


def classify_sesame_rows(conn, rows: Sequence[Dict[str, Any]], stores: Sequence[Any]) -> Dict[str, List[Dict[str, Any]]]:
    by_code = _store_map(stores)
    ext_ids = [r["ext_id"] for r in rows]
    existing = set()
    if ext_ids:
        chunk = 400
        for i in range(0, len(ext_ids), chunk):
            part = ext_ids[i : i + chunk]
            marks = ",".join("?" * len(part))
            existing.update(
                row["ext_id"]
                for row in conn.execute(
                    f"SELECT ext_id FROM advance_posts WHERE ext_id IN ({marks})",
                    part,
                )
            )
    ready: List[Dict[str, Any]] = []
    skipped: List[Dict[str, Any]] = []
    unmatched: List[Dict[str, Any]] = []
    ignored: List[Dict[str, Any]] = []
    for row in rows:
        item = dict(row)
        store = by_code.get(item["mobile_code"])
        if store is not None:
            item["store_id"] = int(store["id"])
            item["store_name"] = store["short_name"] or store["name"]
        else:
            item["store_id"] = None
            item["store_name"] = ""
        if item["status"] and item["status"] != "处理成功":
            item["reason"] = f"状态是{item['status']}"
            ignored.append(item)
            continue
        if not item.get("biz_date"):
            item["reason"] = "没有创建时间"
            ignored.append(item)
            continue
        if not item["mobile_code"] or store is None:
            item["reason"] = "对不上门店编码"
            unmatched.append(item)
            continue
        if item["ext_id"] in existing:
            item["reason"] = "已导入"
            skipped.append(item)
            continue
        if abs(db_advances.yuan_to_cents(item["amount"])) < 1:
            item["reason"] = "金额为 0"
            ignored.append(item)
            continue
        ready.append(item)
    return {"ready": ready, "skipped": skipped, "unmatched": unmatched, "ignored": ignored}


def import_sesame_rows(conn, rows: Sequence[Dict[str, Any]], *, user_id: int) -> int:
    n = 0
    for row in rows:
        biz = row["biz_date"]
        if not isinstance(biz, date):
            biz = parse_created(biz)
        if biz is None:
            raise ValueError(f"流水 {row.get('ext_id') or ''} 没有创建时间")
        note = "芝麻服务费退款" if row.get("refund") else "芝麻服务费"
        order_no = row.get("order_no") or ""
        if order_no:
            note = f"{note} {order_no}"
        db_advances.record_advance(
            conn,
            store_id=int(row["store_id"]),
            user_id=user_id,
            biz_date=biz,
            sesame=row["amount"],
            note=note[:500],
            source="sesame",
            ext_id=row["ext_id"],
            paid=True,
        )
        n += 1
    return n


def _preview_dir() -> Path:
    path = Path(db_core.DATA_DIR) / "tmp" / "sesame"
    path.mkdir(parents=True, exist_ok=True)
    return path


def save_preview(preview: Dict[str, Any]) -> str:
    token = token_hex(16)
    dest = _preview_dir() / f"{token}.json"
    dest.write_text(json.dumps(preview, ensure_ascii=False, default=str), encoding="utf-8")
    _prune_previews()
    return token


def load_preview(token: str) -> Optional[Dict[str, Any]]:
    name = (token or "").strip()
    if not name or any(ch in name for ch in "/\\."):
        return None
    path = _preview_dir() / f"{name}.json"
    if not path.is_file():
        return None
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    return data if isinstance(data, dict) else None


def drop_preview(token: str) -> None:
    name = (token or "").strip()
    if not name or any(ch in name for ch in "/\\."):
        return
    path = _preview_dir() / f"{name}.json"
    try:
        path.unlink(missing_ok=True)
    except OSError:
        pass


def _prune_previews() -> None:
    cutoff = datetime.now(db_core.TZ) - timedelta(hours=2)
    for path in _preview_dir().glob("*.json"):
        try:
            mtime = datetime.fromtimestamp(path.stat().st_mtime, db_core.TZ)
            if mtime < cutoff:
                path.unlink(missing_ok=True)
        except OSError:
            pass


def week_span(day: date) -> tuple[date, date]:
    """自然周：周一到周日。"""
    start = day - timedelta(days=day.weekday())
    return start, start + timedelta(days=6)


def month_span(day: date) -> tuple[date, date]:
    """自然月：1 号到最后一天。"""
    start = day.replace(day=1)
    nxt = date(start.year + (start.month == 12), start.month % 12 + 1, 1)
    return start, nxt - timedelta(days=1)


def month_label(start: date, end: date) -> str:
    """月报标签：跨年用完整日期，同月用「2026年8月」。"""
    if start.year == end.year and start.month == end.month:
        return f"{start.year}年{start.month}月"
    return week_label(start, end)


def period_label(start: date, end: date, mode: str = "week") -> str:
    return month_label(start, end) if mode == "month" else week_label(start, end)


def sesame_tier_breakdown(conn, store_ids, start: date, end: date):
    """每店的档位办理统计，供通报表列。

    返回 (breakdown, tier_cols)：
    - breakdown: {store_id: {"cats": {类别: 办理笔数}, "tiers": {档位: 办理笔数},
                             "tiers_net": {档位: 净额}}}
      办理笔数按净数（扣费 − 退款，实际办理）；tiers 只含「新用户芝麻直降」的原始档位。
    - tier_cols: 本期出现过的直降档位，按数值升序，用作通报表列。
    """
    ids = {int(i) for i in store_ids}
    breakdown: Dict[int, Dict[str, Dict[str, Any]]] = {}
    tiers_all: set = set()
    if not ids:
        return breakdown, []
    for row in db_core.sesame_tier_rows(conn, start.isoformat(), end.isoformat()):
        if row["store_id"] not in ids:
            continue
        d = breakdown.setdefault(
            row["store_id"], {"cats": {}, "tiers": {}, "tiers_net": {}}
        )
        net_n = row["charge_n"] - row["refund_n"]
        d["cats"][row["cat"]] = d["cats"].get(row["cat"], 0) + net_n
        if row["cat"] == "新用户芝麻直降" and row["tier"]:
            d["tiers"][row["tier"]] = d["tiers"].get(row["tier"], 0) + net_n
            d["tiers_net"][row["tier"]] = round(
                d["tiers_net"].get(row["tier"], 0.0) + row["net"], 2
            )
            tiers_all.add(row["tier"])
    tier_cols = sorted(tiers_all, key=lambda t: int(t) if str(t).isdigit() else 9999)
    return breakdown, tier_cols


def week_label(start: date, end: date) -> str:
    if start.year == end.year and start.month == end.month:
        return f"{start.month}月{start.day}日–{end.day}日"
    if start.year == end.year:
        return f"{start.month}月{start.day}日–{end.month}月{end.day}日"
    return f"{start.isoformat()}–{end.isoformat()}"


def sesame_week_rows(
    conn,
    store_ids: Sequence[int],
    start: date,
    end: date,
) -> List[Dict[str, Any]]:
    ids = [int(i) for i in store_ids]
    if not ids:
        return []
    clause, params = db_core.store_in_clause("a.store_id", ids)
    rows = conn.execute(
        f"""
        SELECT a.store_id AS store_id,
               st.name AS name,
               COALESCE(st.city, '') AS city,
               COUNT(*) AS n,
               SUM(CASE WHEN a.sesame > 0 THEN 1 ELSE 0 END) AS charge_n,
               SUM(CASE WHEN a.sesame < 0 THEN 1 ELSE 0 END) AS refund_n,
               ROUND(SUM(CASE WHEN a.sesame > 0 THEN a.sesame ELSE 0 END) / 100.0, 2) AS charge,
               ROUND(SUM(CASE WHEN a.sesame < 0 THEN a.sesame ELSE 0 END) / 100.0, 2) AS refund,
               ROUND(SUM(a.sesame) / 100.0, 2) AS net
        FROM advance_posts a
        JOIN stores st ON st.id = a.store_id
        WHERE a.source='sesame' AND a.biz_date>=? AND a.biz_date<=? AND {clause}
        GROUP BY a.store_id
        ORDER BY net DESC, name
        """,
        [start.isoformat(), end.isoformat(), *params],
    )
    out: List[Dict[str, Any]] = []
    for row in rows:
        charge = float(row["charge"] or 0)
        refund = float(row["refund"] or 0)
        charge_n = int(row["charge_n"] or 0)
        refund_n = int(row["refund_n"] or 0)
        out.append(
            {
                "store_id": int(row["store_id"]),
                "name": row["name"] or "",
                "city": row["city"] or "",
                # 总笔数 = 扣费 − 退款（净笔数）；跨期退款会为负
                "n": charge_n - refund_n,
                "charge_n": charge_n,
                "refund_n": refund_n,
                "charge": charge,
                "refund": refund,
                "refund_abs": round(abs(refund), 2),
                "net": float(row["net"] or 0),
            }
        )
    return out


def sesame_week_totals(rows: Sequence[Mapping[str, Any]]) -> Dict[str, Any]:
    n = charge_n = refund_n = 0
    charge = refund = net = 0.0
    for row in rows:
        n += int(row.get("n") or 0)
        charge_n += int(row.get("charge_n") or 0)
        refund_n += int(row.get("refund_n") or 0)
        charge += float(row.get("charge") or 0)
        refund += float(row.get("refund") or 0)
        net += float(row.get("net") or 0)
    return {
        "n": n,
        "charge_n": charge_n,
        "refund_n": refund_n,
        "charge": round(charge, 2),
        "refund": round(refund, 2),
        "refund_abs": round(abs(refund), 2),
        "net": round(net, 2),
        "stores": len(rows),
    }

def render_week_text(
    rows: Sequence[Mapping[str, Any]],
    totals: Mapping[str, Any],
    start: date,
    end: date,
    city: str = "",
    mode: str = "week",
) -> str:
    city_bit = (city or "").replace("市", "") or "全店"
    title = "芝麻直降办理月报" if mode == "month" else "芝麻直降办理周报"
    lines = [
        f"【{title}】{period_label(start, end, mode)} · {city_bit}",
        f"净 {float(totals.get('net') or 0):.2f} 元 · 净办理 {int(totals.get('n') or 0)} 笔",
        "",
    ]
    if not rows:
        lines.append("这一期还没有已导入的芝麻流水。")
        return "\n".join(lines) + "\n"
    for i, row in enumerate(rows, 1):
        lines.append(
            f"{i} {row.get('name') or '门店'}  净办理{int(row.get('n') or 0)}笔  净{float(row.get('net') or 0):.2f}"
        )
    return "\n".join(lines) + "\n"
