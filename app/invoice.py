"""酬金/租赁开票申请：主体资料 + 按样表导出。"""

from __future__ import annotations

import json
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Mapping, Optional
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from . import db
from .db_invoice import get_invoice_month, month_key
from .helpers import company_names

TEMPLATE = Path(__file__).resolve().parent / "invoice_templates" / "invoice_apply.xlsx"

PARTY_KEYS = ("nt", "tz")
PARTY_FIELDS = (
    "seller_name",
    "seller_tax",
    "seller_addr",
    "seller_bank",
    "buyer_name",
    "buyer_tax",
    "buyer_addr",
    "buyer_bank",
    "email",
)
# 购买方是移动地市公司，样表里已固定；销售方留给设置填执照。
DEFAULT_PARTIES = {
    "nt": {
        "seller_name": "",
        "seller_tax": "",
        "seller_addr": "",
        "seller_bank": "",
        "buyer_name": "中国移动通信集团江苏有限公司南通分公司",
        "buyer_tax": "91320600714057244E",
        "buyer_addr": "江苏省南通市园林路88号0513-68589029",
        "buyer_bank": "中国工商银行股份有限公司南通崇川支行\n1111821209000012463",
        "email": "15162706007@139.com",
    },
    "tz": {
        "seller_name": "",
        "seller_tax": "",
        "seller_addr": "",
        "seller_bank": "",
        "buyer_name": "中国移动通信集团江苏有限公司泰州分公司",
        "buyer_tax": "913212007039775394",
        "buyer_addr": "泰州市医药高新区泰州大道398号   13800523222",
        "buyer_bank": "招商银行北京分行营业部   8888014600006654",
        "email": "15162706007@139.com",
    },
}
_LIMITS = {
    "seller_name": 40,
    "seller_tax": 32,
    "seller_addr": 80,
    "seller_bank": 80,
    "buyer_name": 40,
    "buyer_tax": 32,
    "buyer_addr": 80,
    "buyer_bank": 80,
    "email": 60,
    "handler": 20,
}


def party_key_for_store(store) -> str:
    city = (store["city"] or "").strip()
    return "tz" if "泰州" in city else "nt"


def invoice_store_name(store) -> str:
    raw = ""
    if hasattr(store, "keys") and "invoice_name" in store.keys():
        raw = store["invoice_name"] or ""
    return (raw or "").strip() or (store["name"] or "").strip()


def _clip(value: Any, limit: int) -> str:
    return str(value or "").replace("\r\n", "\n").strip()[:limit]


def default_party(key: str, conn=None) -> Dict[str, str]:
    base = dict(DEFAULT_PARTIES.get(key) or DEFAULT_PARTIES["nt"])
    names = company_names(conn)
    if not base["seller_name"]:
        base["seller_name"] = names.get(key) or ""
    return base


def _load_party(conn, key: str) -> Dict[str, str]:
    out = default_party(key, conn)
    raw = db.get_setting(conn, f"invoice_party_{key}", "")
    if not raw:
        return out
    try:
        loaded = json.loads(raw)
    except (TypeError, ValueError, json.JSONDecodeError):
        return out
    if not isinstance(loaded, dict):
        return out
    for field in PARTY_FIELDS:
        if loaded.get(field):
            out[field] = _clip(loaded.get(field), _LIMITS[field])
    return out


def invoice_parties(conn=None) -> Dict[str, Dict[str, str]]:
    if conn is None:
        with db.get_db() as owned:
            return invoice_parties(owned)
    return {key: _load_party(conn, key) for key in PARTY_KEYS}


def invoice_handler(conn=None) -> str:
    if conn is None:
        with db.get_db() as owned:
            return invoice_handler(owned)
    return _clip(db.get_setting(conn, "invoice_handler", ""), _LIMITS["handler"])


def parse_parties_form(form) -> Dict[str, Dict[str, str]]:
    out = {}
    for key in PARTY_KEYS:
        item = {}
        for field in PARTY_FIELDS:
            item[field] = _clip(form.get(f"{key}_{field}"), _LIMITS[field])
        out[key] = item
    return out


def save_invoice_settings(conn, form) -> None:
    parties = parse_parties_form(form)
    for key, item in parties.items():
        db.set_setting(conn, f"invoice_party_{key}", json.dumps(item, ensure_ascii=False))
    db.set_setting(conn, "invoice_handler", _clip(form.get("handler"), _LIMITS["handler"]))


def _as_date(raw: str, fallback: date) -> date:
    text = (raw or "").strip()[:10]
    if not text:
        return fallback
    try:
        return date.fromisoformat(text)
    except ValueError:
        return fallback


def _set(ws, addr: str, value) -> None:
    ws[addr] = value if value not in ("", None) else None


def _fill_party_block(ws, party: Mapping[str, str]) -> None:
    _set(ws, "B1", party.get("seller_name"))
    _set(ws, "B2", party.get("buyer_name"))
    _set(ws, "B3", party.get("buyer_tax"))
    _set(ws, "B4", party.get("buyer_addr"))
    _set(ws, "B5", party.get("buyer_bank"))


def build_invoice_xlsx(
    conn,
    store,
    as_of: date,
    invoice: Optional[Mapping[str, Any]] = None,
) -> bytes:
    if not TEMPLATE.is_file():
        raise FileNotFoundError("缺少开票申请样表")
    month = month_key(as_of)
    rec = dict(invoice or get_invoice_month(conn, int(store["id"]), month))
    party = _load_party(conn, party_key_for_store(store))
    handler = invoice_handler(conn)
    apply_on = _as_date(rec.get("apply_date") or "", as_of)
    store_title = invoice_store_name(store)
    service = float(rec.get("service") or 0)
    fee = float(rec.get("fee") or 0)
    wb = load_workbook(TEMPLATE)
    comm = wb["酬金开票申请"]
    _fill_party_block(comm, party)
    comm["E9"] = service or None
    comm["E10"] = fee or None
    comm["E9"].number_format = "0.00"
    comm["E10"].number_format = "0.00"
    comm["E11"].number_format = "0.00"
    _set(comm, "B12", store_title)
    _set(comm, "E14", handler)
    comm["E15"] = datetime(apply_on.year, apply_on.month, apply_on.day)
    comm["E15"].number_format = "YYYY/M/D"
    email = party.get("email") or ""
    comm["A25"] = f"接收数电发票的邮箱号：{email}" if email else "接收数电发票的邮箱号："

    detail = wb["明细"]
    detail["B1"] = int(as_of.strftime("%Y%m"))
    detail["C17"] = "=酬金开票申请!E9"
    detail["C17"].number_format = "0.00"
    detail["C28"] = "=酬金开票申请!E10"
    detail["C28"].number_format = "0.00"
    detail["C29"].number_format = "0.00"
    bits = []
    if service:
        bits.append(f"*生产生活服务*服务费{service:.2f}  元")
    if fee:
        bits.append(f"*生产生活服务*手续费 {fee:.2f}  元")
    merged = "D1:D29"
    if merged in [str(r) for r in detail.merged_cells.ranges]:
        detail.unmerge_cells(merged)
    for cell in detail["D1":"D29"]:
        cell[0].value = None
    detail.merge_cells(merged)
    detail["D1"] = "\n".join(bits) if bits else None
    detail["D1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    detail["D1"].font = Font(name="宋体", size=12)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def invoice_filename(store, as_of: date) -> str:
    short = (store["short_name"] or store["name"] or "store").strip()
    return f"invoice_{as_of.strftime('%Y%m')}_{short}.xlsx"


def build_invoice_zip(conn, stores, as_of: date) -> bytes:
    month = month_key(as_of)
    buf = BytesIO()
    with ZipFile(buf, "w", ZIP_DEFLATED) as zf:
        used = set()
        for store in stores:
            name = invoice_filename(store, as_of)
            rec = get_invoice_month(conn, int(store["id"]), month)
            if not rec.get("id"):
                continue  # 无记录的店不占文件名（先判记录再取名）
            if name in used:
                name = f"invoice_{as_of.strftime('%Y%m')}_{store['id']}.xlsx"
            used.add(name)
            rec = get_invoice_month(conn, int(store["id"]), month)
            if not rec.get("id"):
                continue
            data = build_invoice_xlsx(conn, store, as_of, rec)
            zf.writestr(name, data)
    return buf.getvalue()
