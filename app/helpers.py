"""跨路由共享：登录/权限装饰器、门店/日期/播报/考核取数、分页、审计 diff。

原先是 web.py 里 create_app 内的闭包；抽成模块级函数，方便各路由模块复用，
也让 web.py 只负责装配。不含任何 @app.route。
"""

from __future__ import annotations

import re
import time
from datetime import date, timedelta
from functools import wraps
from io import BytesIO
from typing import Any, Dict, List, Optional, Sequence, Tuple

import openpyxl
from flask import Response, current_app, flash, g, redirect, request, session, url_for

from . import broadcast, db, incentive
from .metrics_seed import rollup_amount

DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")

SESSION_IDLE_SECONDS = 30 * 60
UNAUTH_MAX_CONTENT = 64 * 1024

DEFAULT_BRAND = {
    "mark": "vivo",
    "kicker": "零售运营中心",
    "title": "门店日报",
}
BRAND_LIMITS = {"mark": 16, "kicker": 40, "title": 20}


def _clean_brand(value: str, *, default: str, limit: int) -> str:
    text = " ".join((value or "").split())
    if not text:
        return default
    return text[:limit]


def brand_settings(conn=None) -> Dict[str, str]:
    """登录页 / 页头用的标志、副标题、系统名。"""
    if conn is None:
        with db.get_db() as owned:
            return brand_settings(owned)
    hot = db.hot_settings(conn)
    return {
        key: _clean_brand(hot.get(f"brand_{key}", ""), default=default, limit=BRAND_LIMITS[key])
        for key, default in DEFAULT_BRAND.items()
    }


def parse_brand_form(form) -> Dict[str, str]:
    brand = {
        key: _clean_brand(form.get(f"brand_{key}") or "", default="", limit=BRAND_LIMITS[key])
        for key in DEFAULT_BRAND
    }
    if not brand["title"]:
        raise ValueError("系统名称不能空")
    if not brand["mark"]:
        brand["mark"] = DEFAULT_BRAND["mark"]
    return brand


DEFAULT_COMPANY_NAMES = {
    "nt": "南通运营公司",
    "tz": "泰州运营公司",
}
_COMPANY_LIMIT = 30


def company_names(conn=None) -> Dict[str, str]:
    """垫资汇总表按地市归类的两家公司名（南通/泰州）。"""
    if conn is None:
        with db.get_db() as owned:
            return company_names(owned)
    out = {}
    for key, default in DEFAULT_COMPANY_NAMES.items():
        raw = db.get_setting(conn, f"org_name_{key}", "")
        text = " ".join((raw or "").split())[:_COMPANY_LIMIT]
        out[key] = text or default
    return out


def parse_company_names(form) -> Dict[str, str]:
    out = {}
    for key in DEFAULT_COMPANY_NAMES:
        raw = form.get(f"org_name_{key}") or ""
        text = " ".join((raw or "").split())[:_COMPANY_LIMIT]
        out[key] = text or DEFAULT_COMPANY_NAMES[key]
    return out


def review_template_setting(conn=None) -> str:
    """自定义复盘模板；空串表示用内置默认。"""
    if conn is None:
        with db.get_db() as owned:
            return review_template_setting(owned)
    return (db.get_setting(conn, "review_template", "") or "").strip()


def review_preset_key(conn=None) -> str:
    if conn is None:
        with db.get_db() as owned:
            return review_preset_key(owned)
    return (db.get_setting(conn, "review_preset", "") or "").strip()


def load_user() -> None:
    g.user = None
    uid = session.get("user_id")
    if not uid:
        return
    now = time.time()
    last = session.get("_active_at")
    try:
        last_n = float(last) if last is not None else now
    except (TypeError, ValueError):
        last_n = now
    if now - last_n > SESSION_IDLE_SECONDS:
        session.clear()
        return
    with db.get_db() as conn:
        row = conn.execute("SELECT * FROM users WHERE id=? AND active=1", (uid,)).fetchone()
        if row is not None and int(session.get("session_epoch", -1)) != int(row["session_epoch"] or 0):
            # 口令被改过/被重置过：旧会话一律下线
            session.clear()
            return
        g.user = row
        session["_active_at"] = now


def limit_unauth_body():
    """未登录请求体限制在 64KiB，挡住对 /login 的大包拖垮 worker。"""
    if g.user is not None:
        return None
    size = request.content_length or 0
    if size > UNAUTH_MAX_CONTENT:
        return Response("payload too large", status=413)
    return None


def pin_change_required():
    """还是默认口令的账号，只能去改口令或退出。"""
    if g.user is None:
        return None
    keys = g.user.keys()
    if "must_change_pin" not in keys or not int(g.user["must_change_pin"] or 0):
        return None
    endpoint = request.endpoint or ""
    if endpoint in {"settings", "logout", "login", "health", "static"}:
        return None
    flash("请先改掉默认口令，再继续使用", "error")
    return redirect(url_for("settings", tab="account"))


def policy_read_required():
    """开关打开时，未读完启用中的政策不能进其他页。管理员写政策，不拦。"""
    if g.user is None:
        return None
    if g.user["role"] == "admin":
        return None
    endpoint = request.endpoint or ""
    if endpoint in {"policies_page", "policy_ack", "settings", "logout", "login", "health", "static"}:
        return None
    with db.get_db() as conn:
        if not db.policy_require_read(conn):
            return None
        unread = db.unread_policies(conn, g.user["id"])
    if not unread:
        return None
    flash(f"请先阅读并确认 {len(unread)} 条政策，再继续使用。", "error")
    return redirect(url_for("policies_page"))


def csrf_protect():
    """非安全方法必须有合法 CSRF token。测试环境（TESTING）关闭。"""
    if current_app.config.get("TESTING"):
        return None
    if request.method not in ("POST", "PUT", "PATCH", "DELETE"):
        return None
    token = session.get("_csrf_token")
    given = request.form.get("_csrf_token") or request.headers.get("X-CSRF-Token", "")
    if not token or not given or not _compare_digest(token, given):
        # JSON 端点（如编辑器图片上传）拿到 302 HTML 解析不了，回 400 JSON
        if (
            request.accept_mimetypes.best_match(["application/json", "text/html"])
            == "application/json"
        ):
            return Response(
                '{"ok": false, "error": "csrf"}', status=400, mimetype="application/json"
            )
        flash("页面停留太久，操作校验失败，请刷新后重试。", "error")
        # 只跟同源 referrer，防开放重定向。Referer 几乎总是绝对 URL，
        # 同源的转成站内路径再跳；外站、协议相对 //、带反斜杠的一律丢弃。
        target = request.referrer or ""
        if "\\" in target or target.startswith("//"):
            target = ""
        elif target.startswith(request.host_url):
            target = target[len(request.host_url):]  # 同源绝对 URL → 站内路径
            if not target.startswith("/"):
                target = "/" + target
        elif not target.startswith("/"):
            target = ""
        return redirect(target or url_for("today"))
    return None


def _compare_digest(a: str, b: str) -> bool:
    import secrets

    return secrets.compare_digest(a, b)


def default_home(user=None) -> str:
    """只读账号没有填报入口，登录后进报表。"""
    row = user if user is not None else g.user
    if row is not None and row["role"] == "readonly":
        return url_for("report")
    return url_for("today")


def client_ip() -> str:
    """当前请求来源 IP。开了 TRUST_PROXY 时 Flask 已把 X-Forwarded-For 写进 remote_addr。"""
    return (request.remote_addr or "").strip()


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if g.user is None:
            return redirect(url_for("login", next=request.path))
        return fn(*args, **kwargs)

    return wrapper


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if g.user is None:
            return redirect(url_for("login", next=request.path))
        if g.user["role"] != "admin":
            flash("需要管理员权限", "error")
            return redirect(default_home())
        return fn(*args, **kwargs)

    return wrapper


def readonly_required(fn):
    """管理员、只读角色（店长/区域经理）与地市负责人可访问；填报员无权。"""
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if g.user is None:
            return redirect(url_for("login", next=request.path))
        if g.user["role"] not in ("admin", "readonly", "city"):
            flash("需要管理员或只读权限", "error")
            return redirect(default_home())
        return fn(*args, **kwargs)

    return wrapper


def viewer_only(fn):
    """只读与地市负责人不能写（成交/垫资删除等提交），访问页面时直接挡住。"""
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if g.user is not None and g.user["role"] in ("readonly", "city"):
            flash("只读账号不能填写或修改数据", "error")
            return redirect(url_for("report"))
        return fn(*args, **kwargs)

    return wrapper


def store_label(store) -> str:
    if store is None:
        return ""
    short = ""
    try:
        short = (store["short_name"] if "short_name" in store.keys() else "") or ""
    except Exception:
        short = (store.get("short_name") if hasattr(store, "get") else "") or ""
    return short or store["name"]


def broadcast_store_name(store) -> str:
    """群消息只用简称，没有简称才用全称。"""
    return store_label(store)


def parse_date(raw: Optional[str], fallback: Optional[date] = None) -> date:
    if raw and DATE_RE.match(raw):
        try:
            return date.fromisoformat(raw)
        except ValueError:
            pass
    if raw and len(raw) == 7 and raw[4] == "-":
        try:
            return date.fromisoformat(raw + "-01")
        except ValueError:
            pass
    return fallback or db.today_local()


def parse_int(raw, default=0):
    """表单整数：空/非法回落 default，不抛异常。"""
    if raw is None or raw == "":
        return default
    try:
        return int(raw)
    except (TypeError, ValueError):
        return default


def parse_days(raw, default=7, cap=90):
    """天数参数：非法回落 default，再钳到 [1, cap]。"""
    try:
        return max(1, min(int(raw), cap))
    except (TypeError, ValueError):
        return default


def month_end(d: date) -> date:
    """d 所在月的最后一天（12 月跨年）。"""
    if d.month == 12:
        return date(d.year + 1, 1, 1) - timedelta(days=1)
    return date(d.year, d.month + 1, 1) - timedelta(days=1)


def close_rate(closed: int, total: int) -> str:
    """触客成功率：已成交 / 全部触客。没填过显示 —。"""
    closed_n = int(closed or 0)
    total_n = int(total or 0)
    if total_n <= 0:
        return "—"
    # 四舍五入到整数；round() 的银行家舍入会把 32.5% 显示成 32、33.5% 显示成 34
    return f"{int(closed_n * 100 / total_n + 0.5)}%"


def with_close_rate(counts: Dict[str, int]) -> Dict[str, Any]:
    out = dict(counts or {"total": 0, "closed": 0})
    out["rate"] = close_rate(out.get("closed", 0), out.get("total", 0))
    return out


def accessible_stores(conn) -> List[Any]:
    return db.list_user_stores(conn, g.user)


def sql_in(column: str, ids: Sequence[int]) -> Tuple[str, List[int]]:
    clean = [int(i) for i in ids]
    if not clean:
        return "1=0", []
    return f"{column} IN ({','.join('?' * len(clean))})", clean


def scope_options(stores: Sequence[Any]) -> Tuple[List[str], List[str]]:
    cities: List[str] = []
    managers: List[str] = []
    for store in stores:
        city = (store["city"] or "").strip() or "未分地市"
        if city not in cities:
            cities.append(city)
        manager = (store["area_manager"] or "").strip()
        if manager and manager not in managers:
            managers.append(manager)
    return cities, managers


def filter_stores(stores: Sequence[Any], city: str = "", manager: str = "") -> List[Any]:
    city = (city or "").strip()
    manager = (manager or "").strip()
    out: List[Any] = []
    for store in stores:
        store_city = (store["city"] or "").strip() or "未分地市"
        store_manager = (store["area_manager"] or "").strip()
        if city and store_city != city:
            continue
        if manager and store_manager != manager:
            continue
        out.append(store)
    return list(out)


def request_scope(stores: Sequence[Any]) -> Dict[str, Any]:
    """从 ?city= & ?area_manager= 得到当前可见门店。"""
    city = (request.args.get("city") or "").strip()
    manager = (request.args.get("area_manager") or "").strip()
    cities, managers = scope_options(stores)
    if city and city not in cities:
        city = ""
    if manager and manager not in managers:
        manager = ""
    scoped = filter_stores(stores, city, manager)
    bits = [p for p in (city, manager) if p]
    if bits:
        label = " · ".join(bits) + f"（{len(scoped)}家）"
    else:
        label = "全部门店"
    return {
        "city": city,
        "manager": manager,
        "cities": cities,
        "managers": managers,
        "stores": scoped,
        "ids": [int(s["id"]) for s in scoped],
        "label": label,
        "active": bool(city or manager),
    }


def pick_store(conn, raw_id: Optional[str]):
    stores = accessible_stores(conn)
    if not stores:
        return None, []
    if raw_id:
        try:
            sid = int(raw_id)
        except ValueError:
            flash("店号无效，请重新选择门店。", "error")
            return None, stores  # 别静默落到第一家店，避免数据写错店
    else:
        sid = session.get("store_id") or stores[0]["id"]
    if not db.user_can_access_store(conn, g.user, sid):
        sid = stores[0]["id"]
    session["store_id"] = sid
    store = db.get_store(conn, sid)
    return store, stores


def values_for_broadcast(conn, store_id: int, biz_date: date) -> Dict[str, broadcast.DayCum]:
    today = db.day_values(conn, store_id, biz_date)
    prev = db.prev_month_cum(conn, store_id, biz_date)
    return broadcast.add_day_to_prev(prev, today)


def incentive_rules(conn) -> Dict[str, int]:
    return incentive.rules_from(db.get_setting(conn, "incentive_rules", ""))


def broadcast_compact_sections(conn) -> List[str]:
    sections = []
    if db.get_setting(conn, "broadcast_compact", "1") == "1":
        sections.append("digital")
    if db.get_setting(conn, "broadcast_compact_family", "0") == "1":
        sections.append("family")
    return sections


def store_forecast(
    conn,
    store,
    as_of: date,
    rules: Optional[Dict[str, int]] = None,
    *,
    reported: Optional[bool] = None,
    month_vals: Optional[Dict[str, int]] = None,
) -> Dict[str, Any]:
    # month_vals 可传批量预取结果（见 store_forecasts），单店调用不用管
    if month_vals is None:
        month_vals = db.month_cum_through(conn, store["id"], as_of)
    ai = int(month_vals.get("ai_contract", 0) or 0)
    new_cut = rollup_amount(month_vals, "coin_cut")
    advisor_name = (store["advisor_name"] if "advisor_name" in store.keys() else "") or ""
    # rules 默认每次现查；批量循环时调用方应提前算一次传入，避免每店重复查设置
    if rules is None:
        rules = incentive_rules(conn)
    if reported is None:
        start, end = db.month_bounds(as_of)
        reported = (
            conn.execute(
                "SELECT 1 FROM daily_reports WHERE store_id=? AND biz_date>=? AND biz_date<=? LIMIT 1",
                (store["id"], start, end),
            ).fetchone()
            is not None
        )
    if not reported:
        judged = {
            "passed": False,
            "label": "本月未交",
            "reason": "本月还没交过日报，不参与考核",
            "store_reward": 0,
            "store_penalty": 0,
            "advisor_penalty": 0,
            "scheme": "有运营商顾问" if advisor_name.strip() else "无运营商顾问",
            "goal": "先交日报再考核",
            "ai": ai,
            "new_cut": new_cut,
            "sesame": new_cut,
            "total": ai + new_cut,
            "has_advisor": bool(advisor_name.strip()),
            "net": 0,
        }
    else:
        judged = incentive.judge(bool(advisor_name.strip()), ai, new_cut, rules, as_of=as_of)
    judged.update(
        {
            "store_id": store["id"],
            "name": store["name"],
            "store_manager": store["store_manager"] or "",
            "advisor_name": advisor_name.strip(),
            "money_text": incentive.money_text(judged) if reported else "—",
            "reported": bool(reported),
        }
    )
    return judged


def store_forecasts(
    conn, stores, as_of: date, rules: Optional[Dict[str, int]] = None
) -> Dict[int, Dict[str, Any]]:
    """批量考核：月累计、已交日报各一条 SQL 查完，避免结算/考核/看板页每店两条。

    返回 {store_id: judged}。stores 传入前已按权限过滤，结果不出权限范围。
    """
    rules = rules if rules is not None else incentive_rules(conn)
    store_ids = [s["id"] for s in stores]
    reported_ids = set(db.stores_reported_in_month(conn, store_ids, as_of))
    cum = db.month_cum_through_many(conn, store_ids, as_of)
    return {
        s["id"]: store_forecast(
            conn,
            s,
            as_of,
            rules,
            reported=s["id"] in reported_ids,
            month_vals=cum.get(s["id"]),
        )
        for s in stores
    }


def advisor_month_rows(conn, stores, as_of: date, rules: Optional[Dict[str, int]] = None) -> List[Dict[str, Any]]:
    """按顾问聚合当月考核：服务门店、主推奖惩（Σ 名下店的 advisor_penalty）。

    顾问名单来自门店 advisor_name 去重（一人多店合并一行）；
    stores 传入前已按当前用户权限过滤，聚合结果自然不出权限范围。
    """
    rules = rules if rules is not None else incentive_rules(conn)
    divisor = advisor_penalty_divisor(conn)
    judged_map = store_forecasts(conn, stores, as_of, rules)
    by_advisor: Dict[str, Dict[str, Any]] = {}
    for store in stores:
        name = ((store["advisor_name"] if "advisor_name" in store.keys() else "") or "").strip()
        if not name:
            continue
        judged = judged_map[store["id"]]
        item = by_advisor.setdefault(
            name,
            {
                "advisor_name": name,
                "store_ids": [],
                "store_labels": [],
                "penalty_total": 0,
                "store_rows": [],
            },
        )
        label = (store["short_name"] if "short_name" in store.keys() and (store["short_name"] or "").strip() else "") or store["name"]
        item["store_ids"].append(store["id"])
        item["store_labels"].append(label)
        item["penalty_total"] += int(judged.get("advisor_penalty") or 0)
        item["store_rows"].append(
            {
                "store_id": store["id"],
                "label": label,
                "judge_label": judged.get("label") or "",
                "advisor_penalty": int(judged.get("advisor_penalty") or 0),
            }
        )
    rows = list(by_advisor.values())
    for item in rows:
        item["penalty_total"] = round(item["penalty_total"], 2)
        item["fold_coeff"] = round(item["penalty_total"] / divisor, 6)
    rows.sort(key=lambda r: (-r["penalty_total"], r["advisor_name"]))
    return rows


def advisor_penalty_divisor(conn) -> int:
    """主推奖惩折算工资系数的除数，默认 4000（A 类店酬金目标）。"""
    try:
        return max(1, int(db.get_setting(conn, "advisor_penalty_divisor", "4000") or 4000))
    except (TypeError, ValueError):
        return 4000


def advisor_coeffs(base_coeff, fold_coeff, scores) -> Dict[str, Optional[float]]:
    """评分系数 = (店长0.4 + 区域经理0.3 + 地市负责人0.3) / 10；最终 = (基础 + 折算) × 评分。

    与线下「6月运营商顾问工资系数.xlsx」一致：三方都打才出系数。
    """
    if len(scores) != 3 or any(s is None for s in scores):
        return {"rate": None, "final": None}
    sm, sa, sc = (int(s) for s in scores)
    rate = round((sm * 0.4 + sa * 0.3 + sc * 0.3) / 10, 2)
    final = round((float(base_coeff or 1.2) + float(fold_coeff or 0)) * rate, 4)
    return {"rate": rate, "final": final}


def advisor_edit_column(user) -> str:
    """能填的打分列：店长 / 区域经理 / 地市负责人 / 全部；填报员只看不填。"""
    role = user["role"] if user is not None else ""
    if role == "admin":
        return "all"
    if role == "city":
        return "score_city"
    if role == "readonly":
        scope = (user["scope"] or "").strip() if "scope" in user.keys() else ""
        return "score_area" if scope else "score_manager"
    return ""


ADVISOR_SCORE_UNTIL = 5  # 次月 1–5 日给上个月打分


def advisor_score_month(today: date) -> date:
    """默认打分对象：上个月。"""
    return (today.replace(day=1) - timedelta(days=1)).replace(day=1)


def advisor_score_open(today: date, month: date) -> bool:
    """month 是被评月月初。窗口 = 次月 1 日到 5 日。"""
    nxt = month_end(month) + timedelta(days=1)
    return nxt <= today <= nxt.replace(day=ADVISOR_SCORE_UNTIL)


def advisor_score_deadline(month: date) -> date:
    nxt = month_end(month) + timedelta(days=1)
    return nxt.replace(day=ADVISOR_SCORE_UNTIL)


def named_advisor(conn, user) -> str:
    """显示名对上某家启用店的运营商顾问，则返回该顾问名。"""
    name = ""
    if user is not None:
        name = (user["display_name"] or "").strip()
    if not name:
        return ""
    row = conn.execute(
        "SELECT 1 FROM stores WHERE advisor_name=? AND active=1 LIMIT 1",
        (name,),
    ).fetchone()
    return name if row else ""


def build_diff(before: Dict[str, int], after: Dict[str, int], names=None) -> str:
    """把 before/after 值差异拼成可读文本，如「手机销量 1→7」"""
    if names is None:
        from .metrics_seed import metric_name_map
        names = metric_name_map()
    parts = []
    keys = sorted(set(before) | set(after))
    for k in keys:
        b = int(before.get(k, 0) or 0)
        a = int(after.get(k, 0) or 0)
        if b == a:
            continue
        name = names.get(k, k)
        from .metrics_seed import format_stored

        parts.append(f"{name} {format_stored(k, b)}→{format_stored(k, a)}")
    return "；".join(parts) or "（无变化）"


_DEAL_FIELD_LABELS = {
    "closed": "已成交",
    "model": "机型",
    "phone": "号码",
    "spend": "消费",
    "hall_query": "开口",
    "recommend": "推荐套餐",
    "student": "学豆",
    "opener": "导购",
    "note": "备注",
    "text": "播报全文",
}


_DEAL_BOOL_FIELDS = ("closed", "hall_query", "student")


def deal_diff(before: Dict[str, Any], after: Dict[str, Any]) -> str:
    """把成交播报新增/覆盖/删除的差异拼成可读文本。"""
    parts = []
    for key in _DEAL_FIELD_LABELS:
        label = _DEAL_FIELD_LABELS[key]
        b = before.get(key)
        a = after.get(key)
        if b == a:
            continue
        if key in _DEAL_BOOL_FIELDS:
            b_txt = "✓" if b else "✗"
            a_txt = "✓" if a else "✗"
        else:
            b_txt = str(b or "")
            a_txt = str(a or "")
        parts.append(f"{label} {b_txt}→{a_txt}")
    return "；".join(parts) or "（内容未变）"


def pagination(raw_page: Optional[str], total: int, per_page: int = 50) -> Tuple[int, int]:
    """把请求里的 page 参数夹到合法范围，返回 (page, pages)。"""
    try:
        page = max(1, int(raw_page or "1"))
    except (TypeError, ValueError):
        page = 1
    pages = max(1, -(-total // per_page))
    if page > pages:
        page = pages
    return page, pages


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _xlsx_safe(value: Any) -> Any:
    """Prevent formula injection while preserving real numeric cells and templates."""
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        return "'" + value
    return value


def xlsx_bytes(
    header: Sequence[str],
    rows: Sequence[Sequence[Any]],
    *,
    sheet: str = "Sheet1",
    autofilter: bool = True,
) -> bytes:
    """把表头+数据行导出为真实 .xlsx 字节流。表头加粗+底色+冻结，列宽自适应。"""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (sheet or "Sheet1")[:31]  # Excel sheet 名最长 31 字符
    ws.append([_xlsx_safe(value) for value in header])
    fill = PatternFill("solid", fgColor="D9E2F3")
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
    if rows:
        for r in rows:
            ws.append([_xlsx_safe(value) for value in r])
    ws.freeze_panes = "A2"
    if autofilter:
        ws.auto_filter.ref = ws.dimensions
    # 列宽自适应（中文按宽字符估算，封顶 60）
    for col in range(1, max(1, ws.max_column) + 1):
        width = 0.0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                v = cell.value
                w = 0 if v is None else sum(2 if ord(ch) > 127 else 1 for ch in str(v))
                width = max(width, w)
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 10), 60)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def ascii_filename(name: str, fallback: str = "export.xlsx") -> str:
    """下载文件名只留 ASCII，避免部分浏览器把中文文件名拆乱。"""
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", name or "").strip("._")
    return cleaned or fallback


def xlsx_response(data: bytes, filename: str):
    """带下载头的 xlsx 响应。"""
    from flask import Response

    safe = ascii_filename(filename)
    return Response(
        data,
        mimetype=XLSX_MIME,
        headers={"Content-Disposition": f"attachment; filename={safe}"},
    )
