"""运营商月度结算底稿。

系统自动填：区域经理 / 门店 / 类别 / 酬金目标 / AI 目标 / 本月直降 / 本月 AI / 系统奖惩。
留给下月初手填：开票金额、到账房补、垫资、实际搭载率。
公式按已确认口径：
- 实际酬金 = 开票 + 房补 − 垫资
- 酬金得分 = 实际 / 目标，封顶 100
- A 类提成按 4000 档 30/35/40%；B 类按 1000 档同样比例
- 考核奖惩用系统规则（正数奖、负数罚）
- 搭载率目标默认 15%，实际搭载率空着等你另填
"""

from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any, Dict, Iterable, List, Sequence

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .helpers import incentive_rules, store_forecast

COMMISSION_TARGET = {"A": 4000, "B": 1000}
CARRY_TARGET = 0.15
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")


def grade_of(store) -> str:
    raw = ""
    if hasattr(store, "keys") and "store_grade" in store.keys():
        raw = store["store_grade"] or ""
    return (raw or "A").strip().upper()[:1] or "A"


def ai_target_of(store) -> int:
    if hasattr(store, "keys") and "ai_target" in store.keys():
        try:
            return max(0, int(store["ai_target"] or 0))
        except (TypeError, ValueError):
            return 0
    return 0


def commission_target(grade: str) -> int:
    return COMMISSION_TARGET.get(grade, COMMISSION_TARGET["A"])


def focus_score_formula(cell: str, grade: str) -> str:
    """重点业务得分：按本月 AI 台数分档。"""
    if grade == "B":
        return (
            f'IFS({cell}>=4,150,{cell}>=2,100,{cell}>=1,50,{cell}=0,0)'
        )
    return (
        f'IFS({cell}>=10,150,{cell}>=8,120,{cell}>=5,100,{cell}>=3,70,{cell}>=1,60,{cell}=0,0)'
    )


def bonus_formula(actual_cell: str, grade: str) -> str:
    """运营商奖励：A 按 4000，B 按 1000；超额 30/35/40%。"""
    base = commission_target(grade)
    t120 = int(base * 1.2)
    t150 = int(base * 1.5)
    return (
        f"IFS({actual_cell}<{base},0,"
        f"{actual_cell}<={t120},({actual_cell}-{base})*0.3,"
        f"{actual_cell}<={t150},({actual_cell}-{base})*0.35,"
        f"{actual_cell}>{t150},({actual_cell}-{base})*0.4)"
    )


def _style_range(ws, cells, *, fill=None, font=None, align=CENTER, num=None):
    for cell in cells:
        cell.border = THIN
        cell.alignment = align
        if fill is not None:
            cell.fill = fill
        if font is not None:
            cell.font = font
        if num:
            cell.number_format = num


def _group_stores(stores: Sequence[Any]) -> List[Dict[str, Any]]:
    groups: List[Dict[str, Any]] = []
    index: Dict[str, int] = {}
    for store in stores:
        manager = (store["area_manager"] or "").strip() or "未分经理"
        if manager not in index:
            index[manager] = len(groups)
            groups.append({"manager": manager, "stores": []})
        groups[index[manager]]["stores"].append(store)
    return groups


def build_settlement_rows(conn, stores: Iterable[Any], as_of: date) -> List[Dict[str, Any]]:
    rules = incentive_rules(conn)
    out = []
    for store in stores:
        judged = store_forecast(conn, store, as_of, rules)
        grade = grade_of(store)
        out.append(
            {
                "store": store,
                "grade": grade,
                "commission_target": commission_target(grade),
                "ai_target": ai_target_of(store) or (10 if grade == "A" else 2),
                "ai": judged["ai"],
                "new_cut": judged["new_cut"],
                "net": judged["net"],
                "label": judged["label"],
                "money_text": judged["money_text"],
            }
        )
    return out


def build_settlement_xlsx(conn, stores: Sequence[Any], as_of: date) -> bytes:
    rows = build_settlement_rows(conn, stores, as_of)
    by_id = {r["store"]["id"]: r for r in rows}
    groups = _group_stores(stores)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "移动接入"
    title = f"{as_of.month}月通泰零售运营中心运营商绩效"
    ws.merge_cells("A1:S1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    for col in range(1, 20):
        for row in range(2, 5):
            cell = ws.cell(row, col)
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = CENTER
            cell.border = THIN

    ws.merge_cells("A2:A4")
    ws.merge_cells("B2:B4")
    ws.merge_cells("C2:C4")
    ws.merge_cells("D2:D4")
    ws.merge_cells("E2:J2")
    ws.merge_cells("K2:N2")
    ws.merge_cells("O2:Q2")
    ws.merge_cells("R2:R4")
    ws.merge_cells("S2:S4")
    ws["A2"] = "区域经理"
    ws["B2"] = "门店名称"
    ws["C2"] = "门店类别"
    ws["D2"] = "得分"
    ws["E2"] = "酬金（50%）"
    ws["K2"] = "重点业务（30%）"
    ws["O2"] = "业务搭载率（20%）"
    ws["R2"] = "运营商奖励"
    ws["S2"] = "考核奖惩"

    mid = {
        "E3": "酬金目标",
        "F3": "开票金额",
        "G3": "到账房补/调整金额",
        "H3": "垫资",
        "I3": "实际酬金",
        "J3": "酬金得分",
        "K3": "AI手机&直降目标",
        "N3": "得分",
        "O3": "搭载率目标",
        "P3": "实际搭载率",
        "Q3": "得分",
    }
    for coord, text in mid.items():
        ws[coord] = text
    ws.merge_cells("E3:E4")
    ws.merge_cells("F3:F4")
    ws.merge_cells("G3:G4")
    ws.merge_cells("H3:H4")
    ws.merge_cells("I3:I4")
    ws.merge_cells("J3:J4")
    ws.merge_cells("K3:K4")
    ws.merge_cells("L3:M3")
    ws.merge_cells("N3:N4")
    ws.merge_cells("O3:O4")
    ws.merge_cells("P3:P4")
    ws.merge_cells("Q3:Q4")
    ws["L3"] = "完成"
    ws["L4"] = "直降"
    ws["M4"] = "AI手机"

    current = 5
    subtotal_rows: List[int] = []
    all_store_rows: List[int] = []
    for group in groups:
        start = current
        for store in group["stores"]:
            data = by_id[store["id"]]
            _write_store_row(ws, current, group["manager"], data)
            all_store_rows.append(current)
            current += 1
        end = current - 1
        _write_total_row(ws, current, "小计", start, end)
        subtotal_rows.append(current)
        current += 1
    if all_store_rows:
        _write_sum_row(ws, current, "合计", all_store_rows)
        current += 1

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:S{max(4, current - 1)}"
    widths = {
        "A": 12, "B": 36, "C": 8, "D": 8, "E": 12, "F": 12, "G": 16, "H": 10,
        "I": 12, "J": 10, "K": 14, "L": 8, "M": 8, "N": 8, "O": 10, "P": 12,
        "Q": 8, "R": 12, "S": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20

    note = wb.create_sheet("口径")
    note["A1"] = "结算口径（系统生成，勿改公式列）"
    note["A1"].font = Font(bold=True, size=14)
    lines = [
        ("实际酬金", "开票金额 + 到账房补 − 垫资"),
        ("酬金得分", "实际酬金 / 酬金目标 × 100，封顶 100"),
        ("总分", "酬金未满分：酬金50% + 重点30% + 搭载20%；酬金满分则总分 100"),
        ("A 类酬金目标", "4000；提成：未达 0，100%-120% 超额×30%，121%-150%×35%，150%以上×40%"),
        ("B 类酬金目标", "1000；提成档位同比（1200 / 1500）"),
        ("重点业务得分", "按本月 AI 台数分档（A：10/8/5/3/1；B：4/2/1）"),
        ("考核奖惩", "用系统月度考核规则（有顾问看合计，无顾问看双破 0），正数奖、负数罚"),
        ("黄色格子", "下月初手填：开票金额、到账房补、垫资、实际搭载率"),
        ("搭载率", "目标默认 15%，实际搭载率由你另填"),
    ]
    note["A3"] = "项"
    note["B3"] = "规则"
    for i, (k, v) in enumerate(lines, start=4):
        note[f"A{i}"] = k
        note[f"B{i}"] = v
    note.column_dimensions["A"].width = 16
    note.column_dimensions["B"].width = 80

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_store_row(ws, row: int, manager: str, data: Dict[str, Any]) -> None:
    store = data["store"]
    grade = data["grade"]
    ws[f"A{row}"] = manager
    ws[f"B{row}"] = store["name"]
    ws[f"C{row}"] = grade
    ws[f"E{row}"] = data["commission_target"]
    for col in ("F", "G", "H", "P"):
        ws[f"{col}{row}"] = None
        ws[f"{col}{row}"].fill = INPUT_FILL
    ws[f"I{row}"] = f"=F{row}+G{row}-H{row}"
    ws[f"J{row}"] = f'=IF(E{row}=0,0,MIN(100,IF(I{row}/E{row}>=1,100,I{row}/E{row}*100)))'
    ws[f"K{row}"] = data["ai_target"]
    ws[f"L{row}"] = data["new_cut"]
    ws[f"M{row}"] = data["ai"]
    ws[f"N{row}"] = f"={focus_score_formula(f'M{row}', grade)}"
    ws[f"O{row}"] = CARRY_TARGET
    ws[f"Q{row}"] = f'=IF(OR(O{row}=0,P{row}=""),0,P{row}/O{row}*100)'
    ws[f"D{row}"] = (
        f"=MAX(0,MIN(100,IF(J{row}<100,J{row}*0.5+N{row}*0.3+Q{row}*0.2,100)))"
    )
    ws[f"R{row}"] = f"={bonus_formula(f'I{row}', grade)}"
    ws[f"S{row}"] = data["net"]
    ws[f"S{row}"].comment = None
    from openpyxl.comments import Comment

    ws[f"S{row}"].comment = Comment(f"{data['label']} · {data['money_text']}", "store-daily")
    for col in range(1, 20):
        cell = ws.cell(row, col)
        cell.border = THIN
        cell.alignment = LEFT if col == 2 else CENTER
        if col in (6, 7, 8, 16):
            cell.fill = INPUT_FILL
            cell.number_format = "0.00"
        if col in (5, 9, 18, 19):
            cell.number_format = "0.00"
        if col in (4, 10, 14, 17):
            cell.number_format = "0.0"
        if col == 15:
            cell.number_format = "0%"


def _write_total_row(ws, row: int, label: str, start: int, end: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws[f"A{row}"] = label
    for col, letter in enumerate(["E", "F", "G", "H", "I", "K", "L", "M", "R", "S"], start=1):
        ws[f"{letter}{row}"] = f"=SUM({letter}{start}:{letter}{end})"
    for col in range(1, 20):
        cell = ws.cell(row, col)
        cell.border = THIN
        cell.fill = TOTAL_FILL
        cell.font = Font(bold=True)
        cell.alignment = CENTER


def _write_sum_row(ws, row: int, label: str, store_rows: Sequence[int]) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws[f"A{row}"] = label
    for letter in ("E", "F", "G", "H", "I", "K", "L", "M", "R", "S"):
        joined = "+".join(f"{letter}{r}" for r in store_rows)
        ws[f"{letter}{row}"] = "=" + joined
    for col in range(1, 20):
        cell = ws.cell(row, col)
        cell.border = THIN
        cell.fill = TOTAL_FILL
        cell.font = Font(bold=True)
        cell.alignment = CENTER
