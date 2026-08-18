from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from app import db, invoice


def test_save_invoice_and_export_matches_template(app_client):
    app_client.post("/login", data={"username": "admin", "pin": "1234"})
    with db.get_db() as conn:
        store = conn.execute("SELECT * FROM stores WHERE active=1 ORDER BY id LIMIT 1").fetchone()
        sid = int(store["id"])
        db.update_store_profile(
            conn,
            sid,
            mobile_code=store["mobile_code"] or "",
            area_manager=store["area_manager"] or "",
            store_manager=store["store_manager"] or "",
            advisor_name=store["advisor_name"] or "",
            invoice_name="TZ南通市示例体验店",
        )
        invoice.save_invoice_settings(
            conn,
            {
                "handler": "王健",
                "nt_seller_name": "南通财顺电子有限公司",
                "nt_seller_tax": "91320600TEST",
                "nt_seller_addr": "南通测试地址",
                "nt_seller_bank": "测试银行 123",
                "nt_buyer_name": "中国移动通信集团江苏有限公司南通分公司",
                "nt_buyer_tax": "91320600714057244E",
                "nt_buyer_addr": "江苏省南通市园林路88号",
                "nt_buyer_bank": "工行测试账号",
                "nt_email": "15162706007@139.com",
                "tz_seller_name": "泰州市财汇电子有限公司",
                "tz_seller_tax": "",
                "tz_seller_addr": "",
                "tz_seller_bank": "",
                "tz_buyer_name": "中国移动通信集团江苏有限公司泰州分公司",
                "tz_buyer_tax": "913212007039775394",
                "tz_buyer_addr": "",
                "tz_buyer_bank": "",
                "tz_email": "15162706007@139.com",
            },
        )
        rec = db.save_invoice_month(
            conn,
            sid,
            "2026-06",
            service=4854.09,
            fee=1497.5,
            housing=200,
            apply_date="2026-09-02",
        )
        assert rec["invoice_total"] == 6351.59
        store = conn.execute("SELECT * FROM stores WHERE id=?", (sid,)).fetchone()
        data = invoice.build_invoice_xlsx(conn, store, date(2026, 6, 30), rec)

    wb = load_workbook(BytesIO(data))
    assert wb.sheetnames == ["租赁发票开票申请", "酬金开票申请", "明细"]
    comm = wb["酬金开票申请"]
    assert comm["B1"].value == "南通财顺电子有限公司"
    assert comm["B2"].value == "中国移动通信集团江苏有限公司南通分公司"
    assert comm["E9"].value == 4854.09
    assert comm["E10"].value == 1497.5
    assert comm["E11"].value == "=SUM(E9:E10)"
    assert comm["B12"].value == "TZ南通市示例体验店"
    assert comm["E15"].value.month == 9
    detail = wb["明细"]
    assert detail["B1"].value == 202606
    assert detail["B17"].value == "套餐变更"
    assert detail["C17"].value == "=酬金开票申请!E9"
    assert detail["B28"].value == "其他业务"
    assert detail["C28"].value == "=酬金开票申请!E10"
    assert "D1:D29" in [str(r) for r in detail.merged_cells.ranges]
    summary = detail["D1"].value or ""
    assert "*生产生活服务*服务费4854.09  元" in summary
    assert "*生产生活服务*手续费 1497.50  元" in summary
    assert detail["D1"].alignment.horizontal == "center"
    assert detail["D1"].alignment.vertical == "center"
    lease = wb["租赁发票开票申请"]
    assert lease["D9"].value in (None, "")


def test_invoice_page_admin_only_save_and_delete(app_client):
    denied = app_client.get("/incentive/invoice")
    assert denied.status_code in (302, 401, 403)
    app_client.post("/login", data={"username": "alpha", "pin": "123456"})
    filler = app_client.get("/incentive/invoice")
    assert filler.status_code in (302, 403)
    app_client.get("/logout")
    app_client.post("/login", data={"username": "admin", "pin": "1234"})
    page = app_client.get("/incentive/invoice").get_data(as_text=True)
    assert "开票申请" in page
    assert "服务费" in page
    assert "结算月" in page
    assert 'type="month"' in page
    assert "租赁开票" not in page
    with db.get_db() as conn:
        sid = conn.execute("SELECT id FROM stores WHERE active=1 ORDER BY id LIMIT 1").fetchone()["id"]
        db.save_invoice_month(conn, sid, "2026-07", service=10, fee=2)
    august = app_client.get("/incentive?month=2026-08").get_data(as_text=True)
    assert "开票/房补是 2026-07" in august
    assert "month=2026-07" in august
    saved = app_client.post(
        "/incentive/invoice",
        data={
            "month": "2026-06",
            "store_id": str(sid),
            "action": "save",
            "service": "100.5",
            "fee": "20",
            "apply_date": "2026-09-01",
        },
        follow_redirects=True,
    ).get_data(as_text=True)
    assert "开票申请已保存" in saved
    xlsx = app_client.get(f"/incentive/invoice.xlsx?month=2026-06&store_id={sid}")
    assert xlsx.status_code == 200
    deleted = app_client.post(
        "/incentive/invoice",
        data={"month": "2026-06", "store_id": str(sid), "action": "delete"},
        follow_redirects=True,
    ).get_data(as_text=True)
    assert "开票申请已删除" in deleted
    with db.get_db() as conn:
        rec = db.get_invoice_month(conn, sid, "2026-06")
    assert rec["id"] == 0
    settings = app_client.get("/settings?tab=invoice").get_data(as_text=True)
    assert "开票主体" in settings
