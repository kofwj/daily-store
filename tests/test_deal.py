import io

import openpyxl

from app import db
from app.deal import form_values, render_deal
from app.helpers import close_rate, deal_diff, with_close_rate


def test_close_rate_text():
    assert close_rate(0, 0) == "—"
    assert close_rate(1, 2) == "50%"
    assert with_close_rate({"total": 4, "closed": 1})["rate"] == "25%"


def test_closed_deal_uses_short_layout():
    text = render_deal(
        "示例乙店",
        model="S60元气版",
        phone="15514408478",
        spend="99",
        hall_query="1",
        recommend="89",
        closed="1",
        student="0",
        opener="王店长",
        note="用户因价格问题来回拉扯，推荐芝麻直降1600，剩余额度办了分期免息，旧手机抵325元",
    )
    assert text.startswith("示例乙店 · 成交\n")
    assert "S60元气版｜155****8478｜消费99" in text
    assert "15514408478" not in text
    assert "掌厅已查 · 荐89 · 非中高考" in text
    assert "开口 王店长" in text
    assert "🌸" not in text
    assert "是否成交" not in text


def test_open_deal_marks_not_closed():
    text = render_deal("示例甲店", model="X300", closed="0", hall_query="0", student="1")
    assert text.startswith("示例甲店 · 未成交\n")
    assert "未查掌厅" in text
    assert "中高考" in text
    assert "非中高考" not in text


def test_unchecked_boxes_stay_off_after_post():
    values = form_values({"model": "X300"}, posted=True)
    assert values["hall_query"] == "0"
    assert values["closed"] == "0"
    assert values["student"] == "0"
    assert values["show_phone"] == "0"


def test_closed_defaults_off():
    values = form_values()
    assert values["closed"] == "0"
    assert values["hall_query"] == "1"
    assert render_deal("示例乙店").startswith("示例乙店 · 未成交\n")


def test_show_phone_keeps_full_number():
    text = render_deal("示例乙店", phone="15514408478", show_phone="1")
    assert "15514408478" in text


def test_deal_post_counts_by_store(client):
    client.post("/login", data={"username": "admin", "pin": "123456"})
    with db.get_db() as conn:
        sid = conn.execute("SELECT id FROM stores WHERE short_name='示例乙店'").fetchone()["id"]
    client.post(
        "/deal",
        data={"store_id": str(sid), "model": "S60", "phone": "15514408478", "closed": "1", "note": "首发"},
    )
    client.post(
        "/deal",
        data={"store_id": str(sid), "model": "S60", "phone": "15514408478", "closed": "1", "note": "改过"},
    )
    client.post("/deal", data={"store_id": str(sid), "model": "X300", "phone": "13800001111"})
    page = client.get(f"/deal?store_id={sid}").get_data(as_text=True)
    assert "各店次数" not in page  # 记录已迁到独立页
    with db.get_db() as conn:
        today = db.today_local()
        counts = db.deal_counts(conn, [sid], today, today)
        assert counts[sid]["total"] == 2
        assert counts[sid]["closed"] == 1
        rows = list(
            conn.execute(
                "SELECT model, phone, note, text FROM deal_posts WHERE store_id=? ORDER BY id",
                (sid,),
            )
        )
        assert len(rows) == 2
        assert rows[0]["phone"] == "15514408478"
        assert rows[0]["note"] == "改过"
        assert "示例乙店" in rows[0]["text"]
    assert "查看触客记录" in page  # 填报页不再内嵌记录表，只留跳转入口
    rec = client.get(f"/deal/records?store_id={sid}").get_data(as_text=True)
    assert "触客记录" in rec
    assert "S60" in rec
    assert "155****8478" in rec
    assert "15514408478" not in rec
    assert "删除" in rec
    with db.get_db() as conn:
        deal_id = conn.execute(
            "SELECT id FROM deal_posts WHERE store_id=? AND phone='13800001111'",
            (sid,),
        ).fetchone()["id"]
    gone = client.post(
        "/deal/delete",
        data={"store_id": str(sid), "deal_id": str(deal_id)},
        follow_redirects=True,
    ).get_data(as_text=True)
    assert "已删除该触客记录" in gone
    with db.get_db() as conn:
        left = conn.execute("SELECT COUNT(*) AS n FROM deal_posts WHERE store_id=?", (sid,)).fetchone()["n"]
        assert left == 1
    client.get("/logout")
    client.post("/login", data={"username": "beta", "pin": "123456"})
    blocked = client.post(
        "/deal/delete",
        data={"store_id": str(sid), "deal_id": "1"},
        follow_redirects=True,
    )
    assert blocked.status_code == 200
    assert "需要管理员权限" in blocked.get_data(as_text=True)


def test_admin_records_page_shows_all_stores_today(client):
    client.post("/login", data={"username": "admin", "pin": "123456"})
    with db.get_db() as conn:
        sid_a = conn.execute("SELECT id, short_name FROM stores WHERE short_name='示例乙店'").fetchone()
        sid_b = conn.execute("SELECT id, short_name FROM stores WHERE short_name='示例甲店'").fetchone()
        uid = conn.execute("SELECT id FROM users WHERE username='admin'").fetchone()["id"]
        db.record_deal_post(
            conn, store_id=sid_a["id"], user_id=uid, model="S60全店", phone="13811110001", closed=True
        )
        db.record_deal_post(
            conn, store_id=sid_b["id"], user_id=uid, model="X300全店", phone="13822220002", closed=False
        )
    page = client.get("/deal/records").get_data(as_text=True)
    assert "今天全店触客" in page
    assert "S60全店" in page
    assert "X300全店" in page
    assert "示例乙店" in page
    assert "示例甲店" in page
    one = client.get(f"/deal/records?store_id={sid_a['id']}").get_data(as_text=True)
    assert "S60全店" in one
    assert "X300全店" not in one
    client.get("/logout")
    client.post("/login", data={"username": "beta", "pin": "123456"})
    filler = client.get("/deal/records").get_data(as_text=True)
    assert "今天全店触客" not in filler
    assert "全部" not in filler


def test_admin_records_filter_by_city_and_manager(client):
    client.post("/login", data={"username": "admin", "pin": "123456"})
    with db.get_db() as conn:
        nt = conn.execute("SELECT id FROM stores WHERE short_name='示例乙店'").fetchone()["id"]
        tz = conn.execute("SELECT id FROM stores WHERE short_name='示例丁店'").fetchone()["id"]
        uid = conn.execute("SELECT id FROM users WHERE username='admin'").fetchone()["id"]
        db.record_deal_post(conn, store_id=nt, user_id=uid, model="甲店机", phone="13800001111", closed=True)
        db.record_deal_post(conn, store_id=tz, user_id=uid, model="丁店机", phone="13800002222", closed=True)
        mgr = conn.execute("SELECT area_manager FROM stores WHERE id=?", (nt,)).fetchone()["area_manager"]
    page = client.get("/deal/records").get_data(as_text=True)
    assert "全部地市" in page
    assert "全部经理" in page
    assert "甲店机" in page
    assert "丁店机" in page
    city = client.get("/deal/records?city=邻市").get_data(as_text=True)
    assert "丁店机" in city
    assert "甲店机" not in city
    one = client.get(f"/deal/records?area_manager={mgr}").get_data(as_text=True)
    assert "甲店机" in one
    assert "丁店机" not in one


def test_admin_exports_all_stores_deal_csv(client):
    client.post("/login", data={"username": "admin", "pin": "123456"})
    with db.get_db() as conn:
        sid_a = conn.execute("SELECT id FROM stores ORDER BY id LIMIT 1").fetchone()["id"]
        sid_b = conn.execute("SELECT id FROM stores ORDER BY id DESC LIMIT 1").fetchone()["id"]
        uid = conn.execute("SELECT id FROM users WHERE username='admin'").fetchone()["id"]
        db.record_deal_post(
            conn, store_id=sid_a, user_id=uid, model="S60", phone="13811110000",
            spend="99", recommend="89", closed=True, note="甲店",
        )
        db.record_deal_post(
            conn, store_id=sid_b, user_id=uid, model="X300", phone="13822220000",
            spend="199", recommend="139", closed=False, opener="李",
        )
    # 非管理员不能导出
    client.get("/logout")
    client.post("/login", data={"username": "beta", "pin": "123456"})
    r = client.get("/deal/export")
    assert r.status_code == 302
    assert "/today" in r.headers.get("Location", "")
    # 管理员导出全部门店
    client.get("/logout")
    client.post("/login", data={"username": "admin", "pin": "123456"})
    r = client.get("/deal/export")
    assert r.status_code == 200
    assert r.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment" in r.headers.get("Content-Disposition", "")
    assert r.headers.get("Content-Disposition", "").endswith(".xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(r.get_data()))
    ws = wb.active
    assert ws.title == "触客记录"
    header = [c.value for c in ws[1]]
    assert header == ["日期", "门店", "机型", "号码", "消费", "推荐套餐", "开口/导购", "结果", "备注"]
    cells = list(ws.iter_rows(min_row=2, values_only=True))
    flat = [str(v) for row in cells for v in row if v is not None]
    assert "S60" in flat and "13811110000" in flat  # 甲店
    assert "X300" in flat and "13822220000" in flat  # 乙店
    # 只导出设定天数内的（今天必然在内）
    r7 = client.get("/deal/export?days=7")
    assert r7.status_code == 200
    wb7 = openpyxl.load_workbook(io.BytesIO(r7.get_data()))
    flat7 = [str(v) for row in wb7.active.iter_rows(values_only=True) for v in row if v is not None]
    assert "S60" in flat7


def test_record_deal_post_logs_create_and_update(tmp_db):
    """新增写一条 create 审计，覆盖(同店+当日+同号码)写一条带 before 的 update。"""
    import json

    from app import db

    with db.get_db() as conn:
        sid = conn.execute("SELECT id FROM stores WHERE code='store-alpha'").fetchone()["id"]
        uid = conn.execute("SELECT id FROM users WHERE username='admin'").fetchone()["id"]
        # 新增
        did = db.record_deal_post(
            conn,
            store_id=sid,
            user_id=uid,
            closed=True,
            model="S60",
            phone="15500000001",
            spend="99",
            hall_query=True,
            recommend="89",
            opener="张三",
            note="n1",
        )
        rows = list(conn.execute("SELECT * FROM deal_edits WHERE deal_id=?", (did,)))
        assert len(rows) == 1
        assert rows[0]["action"] == "create"
        assert rows[0]["store_id"] == sid
        assert json.loads(rows[0]["after_json"])["model"] == "S60"
        # 覆盖：同店+当日+同号码 -> update
        db.record_deal_post(
            conn,
            store_id=sid,
            user_id=uid,
            closed=True,
            model="S70",
            phone="15500000001",
            spend="199",
            hall_query=True,
            recommend="129",
            opener="张三",
            note="n2",
        )
        rows = list(conn.execute("SELECT * FROM deal_edits ORDER BY id"))
        assert len(rows) == 2
        upd = rows[1]
        assert upd["action"] == "update"
        before = json.loads(upd["before_json"])
        after = json.loads(upd["after_json"])
        assert before["model"] == "S60" and before["spend"] == "99"
        assert after["model"] == "S70" and after["spend"] == "199"
        # 覆盖不会新增 deal_posts 行
        cnt = conn.execute(
            "SELECT COUNT(*) AS c FROM deal_posts WHERE store_id=?", (sid,)
        ).fetchone()["c"]
        assert cnt == 1
        # 删除也写审计
        assert db.delete_deal_post(conn, did, sid, user_id=uid)
        rows = list(conn.execute("SELECT * FROM deal_edits ORDER BY id"))
        assert len(rows) == 3
        assert rows[2]["action"] == "delete"
        assert json.loads(rows[2]["before_json"])["model"] == "S70"
        assert json.loads(rows[2]["after_json"]) == {}


def test_edits_page_filters_by_kind(tmp_db):
    """修改审计里能看到成交播报：按类型筛选 + 全部同时显示。"""
    from app import db
    from app.web import create_app

    app = create_app()
    app.config["TESTING"] = True
    c = app.test_client()
    c.post("/login", data={"username": "admin", "pin": "123456"})
    with db.get_db() as conn:
        sid = conn.execute("SELECT id FROM stores WHERE code='store-alpha'").fetchone()["id"]
        uid = conn.execute("SELECT id FROM users WHERE username='admin'").fetchone()["id"]
        db.record_deal_post(
            conn,
            store_id=sid,
            user_id=uid,
            closed=True,
            model="S60",
            phone="15500000002",
            spend="99",
            hall_query=True,
            recommend="89",
            opener="张三",
            note="n1",
        )
    deal_html = c.get("/edits?kind=deal").get_data(as_text=True)
    assert "成交" in deal_html
    assert "新增触客" in deal_html
    assert "S60" in deal_html
    # 按门店过滤成交
    scoped = c.get(f"/edits?kind=deal&store_id={sid}").get_data(as_text=True)
    assert "新增触客" in scoped
    # 全部默认同时含两种类型（日报造一条）
    page = c.get("/edits").get_data(as_text=True)
    assert "日报" in page and "成交" in page


def test_past_deal_is_readonly(client):
    """当天成交可改可删；往日成交只能看。"""
    from datetime import timedelta

    client.post("/login", data={"username": "admin", "pin": "123456"})
    with db.get_db() as conn:
        sid = conn.execute("SELECT id FROM stores WHERE short_name='示例乙店'").fetchone()["id"]
        uid = conn.execute("SELECT id FROM users WHERE username='admin'").fetchone()["id"]
        today = db.today_local()
        yesterday = today - timedelta(days=1)
        past_id = db.record_deal_post(
            conn,
            store_id=sid,
            user_id=uid,
            closed=True,
            model="昨日机",
            phone="13900001111",
            note="昨天的",
            biz_date=yesterday,
        )
        today_id = db.record_deal_post(
            conn,
            store_id=sid,
            user_id=uid,
            closed=True,
            model="今日机",
            phone="13900002222",
            note="今天的",
            biz_date=today,
        )
    past_page = client.get(f"/deal?store_id={sid}&deal_id={past_id}").get_data(as_text=True)
    assert "往日触客只能查看" in past_page
    assert "昨日机" in past_page
    assert 'name="model"' not in past_page
    assert "生成播报" not in past_page
    today_page = client.get(f"/deal?store_id={sid}&deal_id={today_id}").get_data(as_text=True)
    assert "生成播报" in today_page
    assert "今日机" in today_page
    blocked = client.post(
        "/deal",
        data={
            "store_id": str(sid),
            "deal_id": str(past_id),
            "model": "改不了",
            "phone": "13900001111",
            "note": "不该写进去",
        },
        follow_redirects=True,
    ).get_data(as_text=True)
    assert "往日触客只能查看，不能改" in blocked
    rec = client.get(f"/deal/records?store_id={sid}&days=7").get_data(as_text=True)
    assert "只读" in rec
    gone = client.post(
        "/deal/delete",
        data={"store_id": str(sid), "deal_id": str(past_id)},
        follow_redirects=True,
    ).get_data(as_text=True)
    assert "往日触客只能查看，不能删" in gone
    with db.get_db() as conn:
        still = conn.execute("SELECT model, note FROM deal_posts WHERE id=?", (past_id,)).fetchone()
        assert still["model"] == "昨日机"
        assert still["note"] == "昨天的"
        assert db.delete_deal_post(conn, today_id, sid, user_id=uid)
        assert db.delete_deal_post(conn, past_id, sid, user_id=uid) is False


def test_new_deal_defaults_opener_to_store_advisor(client):
    client.post("/login", data={"username": "admin", "pin": "123456"})
    with db.get_db() as conn:
        sid = conn.execute("SELECT id FROM stores WHERE short_name='示例乙店'").fetchone()["id"]
        conn.execute("UPDATE stores SET advisor_name='王店长' WHERE id=?", (sid,))
    page = client.get(f"/deal?store_id={sid}").get_data(as_text=True)
    assert 'name="opener"' in page
    assert 'value="王店长"' in page


def test_deal_diff_formats_bool_fields():
    text = deal_diff(
        {"closed": False, "hall_query": True, "student": False, "model": "S60"},
        {"closed": True, "hall_query": False, "student": True, "model": "S70"},
    )
    assert "已成交 ✗→✓" in text
    assert "开口 ✓→✗" in text
    assert "学豆 ✗→✓" in text
    assert "机型 S60→S70" in text


def test_deal_unknown_id_raises_not_overwrite(tmp_db):
    """无效 deal_id 不能静默落去重分支改别人的记录。"""
    import pytest

    from app.deal import is_today_deal  # noqa: F401 — 借用触发 app 导入
    with db.get_db() as conn:
        sid = conn.execute("SELECT id FROM stores WHERE code='store-alpha'").fetchone()["id"]
        with pytest.raises(ValueError):
            db.record_deal_post(
                conn,
                store_id=sid,
                user_id=1,
                closed=True,
                model="X300",
                phone="13900000000",
                deal_id=999999,
                biz_date=db.today_local(),
            )
