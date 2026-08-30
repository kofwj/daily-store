from datetime import date
from io import BytesIO

import openpyxl

from app import db
from app.stores_seed import STORES


def test_catalog_includes_advance_workbook_stores(tmp_db):
    names = {item["name"] for item in STORES}
    assert "示例市甲街vivo体验店" in names
    assert "邻市戊路vivo体验店" in names
    assert "邻市巳街vivo专卖店" in names
    with db.get_db() as conn:
        row = conn.execute("SELECT * FROM stores WHERE code='store-gamma'").fetchone()
        assert row is not None
        assert row["area_manager"] == "张管理"


def test_filler_must_provide_phone(client):
    client.post("/login", data={"username": "alpha", "pin": "123456"})
    with db.get_db() as conn:
        sid = conn.execute("SELECT id FROM stores WHERE code='store-alpha'").fetchone()["id"]
    page = client.post(
        "/advance",
        data={"store_id": str(sid), "biz_date": db.today_local().isoformat(), "rebate": "100", "note": "购机让利"},
        follow_redirects=True,
    ).get_data(as_text=True)
    assert "必须带号码" in page
    with db.get_db() as conn:
        n = conn.execute("SELECT COUNT(*) AS n FROM advance_posts WHERE store_id=?", (sid,)).fetchone()["n"]
        assert n == 0


def test_filler_saves_and_admin_pays(client):
    client.post("/login", data={"username": "alpha", "pin": "123456"})
    with db.get_db() as conn:
        sid = conn.execute("SELECT id FROM stores WHERE code='store-alpha'").fetchone()["id"]
    today = db.today_local().isoformat()
    saved = client.post(
        "/advance",
        data={
            "store_id": str(sid),
            "biz_date": today,
            "phone": "13900001111",
            "broadband": "200",
            "rebate": "100",
            "note": "宽带电视",
        },
        follow_redirects=True,
    ).get_data(as_text=True)
    assert "垫资已保存" in saved
    # 店员视角号码打码
    assert "139****1111" in saved
    assert "13900001111" not in saved
    assert "未兑" in saved
    with db.get_db() as conn:
        aid = conn.execute("SELECT id FROM advance_posts WHERE store_id=?", (sid,)).fetchone()["id"]
    client.post("/logout")
    client.post("/login", data={"username": "admin", "pin": "123456"})
    paid = client.post(
        "/advance/pay",
        data={"action": "pay", "advance_id": [str(aid)], "month": today[:7], "paid": "0"},
        follow_redirects=True,
    ).get_data(as_text=True)
    assert "已兑付 1 笔" in paid
    locked = client.post(
        "/advance",
        data={
            "store_id": str(sid),
            "advance_id": str(aid),
            "biz_date": today,
            "phone": "13900001111",
            "rebate": "50",
        },
        follow_redirects=True,
    ).get_data(as_text=True)
    assert "已兑付的垫资不能改" in locked
    gone = client.post(
        "/advance/delete",
        data={"store_id": str(sid), "advance_id": str(aid)},
        follow_redirects=True,
    ).get_data(as_text=True)
    assert "已兑付的垫资不能删" in gone
    with db.get_db() as conn:
        row = conn.execute("SELECT paid, rebate, broadband FROM advance_posts WHERE id=?", (aid,)).fetchone()
        assert int(row["paid"]) == 1
        assert int(row["broadband"]) == 20000
        assert int(row["rebate"]) == 10000


def test_admin_can_save_without_phone_and_fills_settlement(tmp_db, admin_client):
    c = admin_client
    with db.get_db() as conn:
        sid = conn.execute("SELECT id FROM stores WHERE code='store-alpha'").fetchone()["id"]
    today = db.today_local()
    page = c.post(
        "/advance",
        data={
            "store_id": str(sid),
            "biz_date": today.isoformat(),
            "other": "39.99",
            "note": "芝麻服务费",
        },
        follow_redirects=True,
    ).get_data(as_text=True)
    assert "垫资已保存" in page
    assert "芝麻服务费" in page
    r = c.get(f"/incentive.xlsx?month={today.strftime('%Y-%m')}")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(BytesIO(r.get_data()))
    ws = wb["移动接入"]
    found = False
    for row in ws.iter_rows(min_row=5, max_col=8, values_only=True):
        if row[1] == "示例市甲街vivo体验店":
            assert row[7] == 39.99
            found = True
            break
    assert found
    export = c.get(f"/advance.xlsx?month={today.strftime('%Y-%m')}")
    assert export.status_code == 200
    book = openpyxl.load_workbook(BytesIO(export.get_data()))
    assert "汇总表" in book.sheetnames
    assert "示例甲店" in book.sheetnames
    assert book["示例甲店"]["H3"].value == "芝麻服务费"
    names = [row[0] for row in book["汇总表"].iter_rows(min_col=1, max_col=1, values_only=True)]
    assert "示例公司甲" not in names
    assert "南通运营公司" in names or any("南通" in str(n) for n in names if n)


def test_filler_rejects_future_and_nonfinite_amounts(client):
    from datetime import timedelta

    client.post("/login", data={"username": "alpha", "pin": "123456"})
    with db.get_db() as conn:
        sid = conn.execute("SELECT id FROM stores WHERE code='store-alpha'").fetchone()["id"]
    future = client.post("/advance", data={"store_id": sid, "biz_date": (db.today_local() + timedelta(days=1)).isoformat(), "phone": "13900000000", "rebate": "10"}, follow_redirects=True)
    assert "未来日期" in future.get_data(as_text=True)
    invalid = client.post("/advance", data={"store_id": sid, "biz_date": db.today_local().isoformat(), "phone": "13900000000", "rebate": "nan"}, follow_redirects=True)
    assert "金额请填数字" in invalid.get_data(as_text=True)


def test_filler_can_save_negative_amount(client):
    client.post("/login", data={"username": "alpha", "pin": "123456"})
    with db.get_db() as conn:
        sid = conn.execute("SELECT id FROM stores WHERE code='store-alpha'").fetchone()["id"]
    page = client.post(
        "/advance",
        data={
            "store_id": str(sid),
            "biz_date": db.today_local().isoformat(),
            "phone": "13900004444",
            "broadband": "-100",
            "note": "退网络电视",
        },
        follow_redirects=True,
    ).get_data(as_text=True)
    assert "垫资已保存" in page
    assert "退款填负数" in page
    with db.get_db() as conn:
        row = conn.execute(
            "SELECT broadband FROM advance_posts WHERE store_id=? AND phone='13900004444'",
            (sid,),
        ).fetchone()
        assert int(row["broadband"]) == -10000


def test_advance_actions_are_audited(client):
    client.post("/login", data={"username": "alpha", "pin": "123456"})
    with db.get_db() as conn:
        sid = conn.execute("SELECT id FROM stores WHERE code='store-alpha'").fetchone()["id"]
    today = db.today_local().isoformat()
    client.post("/advance", data={"store_id": sid, "biz_date": today, "phone": "13900007777", "rebate": "10"})
    with db.get_db() as conn:
        aid = conn.execute("SELECT id FROM advance_posts WHERE phone='13900007777'").fetchone()["id"]
    client.post("/advance", data={"store_id": sid, "advance_id": aid, "biz_date": today, "phone": "13900007777", "rebate": "20"})
    client.post("/logout")
    client.post("/login", data={"username": "admin", "pin": "123456"})
    client.post("/advance/pay", data={"action": "pay", "advance_id": [str(aid)]})
    client.post("/advance/pay", data={"action": "unpay", "advance_id": [str(aid)]})
    client.post("/advance/delete", data={"store_id": sid, "advance_id": aid})
    with db.get_db() as conn:
        actions = [row["action"] for row in conn.execute("SELECT action FROM advance_edits ORDER BY id")]
    assert actions == ["create", "update", "pay", "unpay", "delete"]


def test_anonymous_cannot_delete_advance(client):
    client.post("/login", data={"username": "alpha", "pin": "123456"})
    with db.get_db() as conn:
        sid = conn.execute("SELECT id FROM stores WHERE code='store-alpha'").fetchone()["id"]
    today = db.today_local().isoformat()
    client.post(
        "/advance",
        data={
            "store_id": str(sid),
            "biz_date": today,
            "phone": "13900006666",
            "rebate": "10",
        },
        follow_redirects=True,
    )
    with db.get_db() as conn:
        aid = conn.execute(
            "SELECT id FROM advance_posts WHERE phone='13900006666'"
        ).fetchone()["id"]
    client.post("/logout")
    resp = client.post(
        "/advance/delete",
        data={"store_id": str(sid), "advance_id": str(aid)},
        follow_redirects=False,
    )
    assert resp.status_code == 302
    assert "/login" in (resp.headers.get("Location") or "")
    with db.get_db() as conn:
        n = conn.execute(
            "SELECT COUNT(*) AS n FROM advance_posts WHERE id=?", (aid,)
        ).fetchone()["n"]
        assert n == 1


def test_filler_cannot_open_pay_page(client):
    client.post("/login", data={"username": "alpha", "pin": "123456"})
    r = client.get("/advance/pay", follow_redirects=True)
    assert "需要管理员权限" in r.get_data(as_text=True)


def test_admin_advance_defaults_to_all_stores(client):
    client.post("/login", data={"username": "alpha", "pin": "123456"})
    with db.get_db() as conn:
        sid_a = conn.execute("SELECT id FROM stores WHERE code='store-alpha'").fetchone()["id"]
        sid_b = conn.execute("SELECT id FROM stores WHERE code='store-beta'").fetchone()["id"]
    today = db.today_local().isoformat()
    client.post(
        "/advance",
        data={"store_id": str(sid_a), "biz_date": today, "phone": "13900009111", "rebate": "10"},
    )
    client.post("/logout")
    client.post("/login", data={"username": "beta", "pin": "123456"})
    client.post(
        "/advance",
        data={"store_id": str(sid_b), "biz_date": today, "phone": "13900009222", "rebate": "20"},
    )
    client.post("/logout")
    client.post("/login", data={"username": "admin", "pin": "123456"})
    page = client.get("/advance").get_data(as_text=True)
    assert "本月全店垫资" in page
    assert "13900009111" in page
    assert "13900009222" in page
    assert "示例甲店" in page
    assert "示例乙店" in page
    assert "记一笔垫资" not in page
    one = client.get(f"/advance?store_id={sid_a}").get_data(as_text=True)
    assert "记一笔垫资" in one
    assert "13900009111" in one
    assert "13900009222" not in one


def test_admin_advance_filter_by_city(client):
    client.post("/login", data={"username": "admin", "pin": "123456"})
    with db.get_db() as conn:
        nt = conn.execute("SELECT id FROM stores WHERE short_name='示例乙店'").fetchone()["id"]
        tz = conn.execute("SELECT id FROM stores WHERE short_name='示例丁店'").fetchone()["id"]
    today = db.today_local().isoformat()
    client.post("/advance", data={"store_id": str(nt), "biz_date": today, "phone": "13900009333", "rebate": "11"})
    client.post("/advance", data={"store_id": str(tz), "biz_date": today, "phone": "13900009444", "rebate": "22"})
    all_page = client.get("/advance").get_data(as_text=True)
    assert "13900009333" in all_page
    assert "13900009444" in all_page
    city = client.get("/advance?city=邻市").get_data(as_text=True)
    assert "13900009444" in city
    assert "13900009333" not in city


def test_store_sees_month_list_and_admin_sees_today_inbox(client):
    client.post("/login", data={"username": "alpha", "pin": "123456"})
    with db.get_db() as conn:
        sid = conn.execute("SELECT id FROM stores WHERE code='store-alpha'").fetchone()["id"]
    today = db.today_local().isoformat()
    client.post(
        "/advance",
        data={
            "store_id": str(sid),
            "biz_date": today,
            "phone": "13900003333",
            "rebate": "100",
            "note": "购机让利",
        },
        follow_redirects=True,
    )
    page = client.get("/advance").get_data(as_text=True)
    assert "本月记录" in page
    # 店员视角号码打码
    assert "139****3333" in page
    assert "13900003333" not in page
    assert "改" in page and "删" in page
    client.post("/logout")
    client.post("/login", data={"username": "admin", "pin": "123456"})
    inbox = client.get("/advance/pay?scope=today&paid=0").get_data(as_text=True)
    assert "今天待兑" in inbox
    assert "示例甲店" in inbox
    assert "1笔" in inbox


def test_advance_stores_cents_and_reads_yuan(client):
    client.post("/login", data={"username": "alpha", "pin": "123456"})
    with db.get_db() as conn:
        sid = conn.execute("SELECT id FROM stores WHERE code='store-alpha'").fetchone()["id"]
    client.post(
        "/advance",
        data={
            "store_id": str(sid),
            "biz_date": db.today_local().isoformat(),
            "phone": "13900008888",
            "other": "39.99",
        },
        follow_redirects=True,
    )
    with db.get_db() as conn:
        raw = conn.execute(
            "SELECT other FROM advance_posts WHERE phone='13900008888'"
        ).fetchone()
        viewed = db.get_advance(conn, conn.execute(
            "SELECT id FROM advance_posts WHERE phone='13900008888'"
        ).fetchone()["id"], sid)
    assert int(raw["other"]) == 3999
    assert float(viewed["other"]) == 39.99
    assert float(viewed["total"]) == 39.99


def test_advance_range_sums_and_pay_page(client):
    client.post("/login", data={"username": "admin", "pin": "123456"})
    with db.get_db() as conn:
        sid = conn.execute("SELECT id FROM stores WHERE code='store-alpha'").fetchone()["id"]
    today = db.today_local()
    client.post(
        "/advance",
        data={"store_id": str(sid), "biz_date": today.isoformat(), "phone": "13900009555", "rebate": "10.5"},
    )
    client.post(
        "/advance",
        data={"store_id": str(sid), "biz_date": today.isoformat(), "phone": "13900009556", "broadband": "20"},
    )
    with db.get_db() as conn:
        aid = conn.execute("SELECT id FROM advance_posts WHERE phone='13900009555'").fetchone()["id"]
        sums = db.advance_range_sums(conn, start=today, end=today, store_id=sid)
    assert sums["rebate"] == 10.5
    assert sums["broadband"] == 20.0
    assert sums["total"] == 30.5
    assert sums["unpaid"] == 2
    client.post("/advance/pay", data={"action": "pay", "advance_id": [str(aid)]})
    page = client.get("/advance/pay?scope=today").get_data(as_text=True)
    assert "30.50" in page or "30.5" in page
    assert "未兑 1 笔" in page
    with db.get_db() as conn:
        after = db.advance_range_sums(conn, start=today, end=today, store_id=sid)
    assert after["unpaid"] == 1
    assert after["total"] == 30.5


def test_advance_old_real_whole_yuan_converted(tmp_path):
    """旧库 REAL 列存的是元——整元（200、500）也必须乘100，不能只靠采样猜。"""

    import sqlite3

    from app import db_core

    conn = sqlite3.connect(str(tmp_path / "old.db"))
    db_core._ensure_app_meta(conn)
    conn.execute("CREATE TABLE advance_posts (id INTEGER PRIMARY KEY AUTOINCREMENT, broadband REAL NOT NULL DEFAULT 0, rebate REAL NOT NULL DEFAULT 0, other REAL NOT NULL DEFAULT 0)")
    conn.executemany(
        "INSERT INTO advance_posts (broadband, rebate, other) VALUES (?, ?, ?)",
        [(200, 35.5, 0), (500, 0, 12.25)],
    )
    db_core._advance_amounts_to_cents(conn)
    rows = conn.execute(
        "SELECT broadband, rebate, other FROM advance_posts ORDER BY id"
    ).fetchall()
    # 整元 200→20000 分、500→50000 分；小数照转
    assert list(rows) == [(20000, 3550, 0), (50000, 0, 1225)]
    marker = conn.execute("SELECT value FROM app_meta WHERE key='advance_cents_marker'").fetchone()
    assert marker[0] == "1"
    conn.close()


def test_advance_form_keeps_input_on_error(filler_client):
    """垫资校验失败要回显已填内容，不能让店员重敲一遍。"""
    with db.get_db() as conn:
        sid = conn.execute("SELECT id FROM stores WHERE code='store-alpha'").fetchone()["id"]
    resp = filler_client.post(
        "/advance",
        data={
            "store_id": str(sid),
            "biz_date": date.today().isoformat(),
            "phone": "13800138000",
            "broadband": "abc",
            "note": "宽带垫资备注",
        },
    )
    page = resp.get_data(as_text=True)
    assert "金额请填数字" in page
    assert "13800138000" in page
    assert "宽带垫资备注" in page


def test_advance_phone_masked_for_filler(filler_client):
    """垫资记录对店员打码，管理员保留完整号码（兑付对账用）。"""
    with db.get_db() as conn:
        sid = conn.execute("SELECT id FROM stores WHERE code='store-alpha'").fetchone()["id"]
        admin_id = conn.execute("SELECT id FROM users WHERE username='admin'").fetchone()["id"]
        db.record_advance(
            conn, store_id=sid, user_id=admin_id, biz_date=date.today(),
            phone="13812345678", broadband=100,
        )
    filler_page = filler_client.get("/advance").get_data(as_text=True)
    assert "138****5678" in filler_page
    assert "13812345678" not in filler_page
    filler_client.post("/logout")
    filler_client.post("/login", data={"username": "admin", "pin": "123456"})
    admin_page = filler_client.get("/advance").get_data(as_text=True)
    assert "13812345678" in admin_page
