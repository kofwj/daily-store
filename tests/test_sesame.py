"""芝麻服务费导入：解析、对店、去重、垫资联动、店员只读。"""

from datetime import date
from io import BytesIO

import openpyxl
from openpyxl import Workbook

from app import db

# 样本数据业务日：跟随机器当前月，避免跨月后「本月视图 / 本月合计」断言失配。
# 取本月 15 号（月初则取今天，保证 ≤ 今天且在本月内）。
_REF = date.today().replace(day=min(15, date.today().day))
_REF_ISO = _REF.isoformat()
_REF_MONTH = _REF.strftime("%Y%m")
_MONTH_START = _REF.replace(day=1).isoformat()


def _make_sesame_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "芝麻服务费明细1"
    ws.append(["芝麻服务费明细"])
    headers = [
        "流水号", "订单号", "省份", "城市", "地区", "商户号", "营业执照号",
        "营业执照名称", "门店编码", "门店名称", "门店类型", "统计月份",
        "订单金额", "服务费金额", "状态", "备注", "创建时间",
    ]
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sample_rows():
    return [
        [
            "1786774661473466851", "1786774661473466851", "江苏省", "示例市", "甲区",
            "JSCM_20250116110103117937984", "91320602MA7E6JC70A", "示例公司甲",
            "JSCM_10000001", "示例甲店vivo专营店", "加盟店", _REF_MONTH,
            1368.00, -9.58, "处理成功",
            "行业芝麻订单(1786774661473466851)服务费", f"{_REF_ISO} 14:19:38.0",
        ],
        [
            "1786527145403143969_R", "1786527145403143969", "江苏省", "示例市", "乙区",
            "JSCM_20250116110103117937984", "91320602MA7E6JC70A", "示例公司甲",
            "JSCM_10000004", "示例戊店vivo专卖店", "加盟店", _REF_MONTH,
            1368.00, 9.58, "处理成功",
            "行业芝麻订单(1786527145403143969)服务费退款", f"{_REF_ISO} 09:22:26.0",
        ],
        [
            "1785792754901613702", "1785792754901613702", "江苏省", "邻市", "丙区",
            "JSCM_20250116120104683937984", "91321202MA7M941H5H", "示例公司乙",
            "JSCM_10000003", "示例戊店带店vivo专卖店", "加盟店", _REF_MONTH,
            792.00, -5.54, "处理成功",
            "行业芝麻订单(1785792754901613702)服务费", f"{_REF_ISO} 19:21:56.0",
        ],
        [
            "9999999999999999999", "9999999999999999999", "江苏省", "示例市", "丁区",
            "JSCM_20250116110103117937984", "91320602MA7E6JC70A", "示例公司甲",
            "JSCM_99999999", "对不上的店", "加盟店", _REF_MONTH,
            480.00, -3.36, "处理成功",
            "行业芝麻订单(9999999999999999999)服务费", f"{_REF_ISO} 10:00:00.0",
        ],
    ]


def test_parse_official_sesame_workbook():
    from pathlib import Path

    path = Path(__file__).resolve().parent.parent / "芝麻服务费明细20260817082042.xlsx"
    if not path.is_file():
        return
    rows = db.parse_sesame_xlsx(path.read_bytes())
    assert len(rows) >= 40
    assert rows[0]["mobile_code"]
    assert isinstance(rows[0]["amount"], float)


def test_parse_sesame_xlsx(tmp_db):
    data = _make_sesame_xlsx(_sample_rows())
    rows = db.parse_sesame_xlsx(data)
    assert len(rows) == 4
    r0 = rows[0]
    assert r0["mobile_code"] == "10000001"
    assert r0["amount"] == 9.58  # 取反
    assert r0["refund"] is False
    assert r0["biz_date"] == _REF_ISO
    r1 = rows[1]
    assert r1["amount"] == -9.58  # 退款取反为负
    assert r1["refund"] is True
    r2 = rows[2]
    assert r2["mobile_code"] == "10000003"  # 示例丁店


def test_classify_matches_jiangyan_and_xinghua(tmp_db):
    data = _make_sesame_xlsx(_sample_rows())
    rows = db.parse_sesame_xlsx(data)
    with db.get_db() as conn:
        stores = list(conn.execute("SELECT * FROM stores WHERE active=1"))
        groups = db.classify_sesame_rows(conn, rows, stores)
    codes = {r["mobile_code"] for r in groups["ready"]}
    assert "10000003" in codes  # 示例丁店
    unmatched_codes = {r["mobile_code"] for r in groups["unmatched"]}
    assert "99999999" in unmatched_codes


def test_import_creates_advance_rows_and_dedup(tmp_db, admin_client):
    c = admin_client
    data = _make_sesame_xlsx(_sample_rows())
    # 第一次预览
    c.post("/advance/sesame/preview", data={"sesame_file": (BytesIO(data), "sesame.xlsx")})
    # 预览存 session，需要先拿到 preview 页确认
    page = c.get("/advance/sesame").get_data(as_text=True)
    assert "可导入" in page
    assert "示例甲店" in page
    assert "9.58" in page
    assert "示例丁店" in page
    assert "对不上门店" in page
    # 确认导入
    confirm = c.post("/advance/sesame/confirm", follow_redirects=True).get_data(as_text=True)
    assert "已导入 3 笔" in confirm
    # 库里检查
    with db.get_db() as conn:
        rows = list(conn.execute(
            "SELECT sesame, source, ext_id, note, paid FROM advance_posts WHERE source='sesame' ORDER BY ext_id"
        ))
    assert len(rows) == 3
    assert all(int(r["paid"] or 0) == 1 for r in rows)
    amounts = sorted(float(r["sesame"]) / 100 for r in rows)
    assert amounts == [-9.58, 5.54, 9.58]
    # 再次预览：已导入的应跳过
    c.post("/advance/sesame/preview", data={"sesame_file": (BytesIO(data), "sesame.xlsx")})
    page2 = c.get("/advance/sesame").get_data(as_text=True)
    assert "已导入跳过" in page2
    assert "可导入" not in page2 or "0" in page2


def test_imported_rows_locked_from_filler(tmp_db, admin_client):
    c = admin_client
    data = _make_sesame_xlsx(_sample_rows()[:1])
    c.post("/advance/sesame/preview", data={"sesame_file": (BytesIO(data), "s.xlsx")})
    c.post("/advance/sesame/confirm", follow_redirects=True)
    c.post("/logout")
    # 店员登录
    c.post("/login", data={"username": "alpha", "pin": "123456"})
    with db.get_db() as conn:
        row = conn.execute(
            "SELECT id, store_id FROM advance_posts WHERE source='sesame' LIMIT 1"
        ).fetchone()
    assert row is not None
    # 店员不能改
    edit = c.post(
        "/advance",
        data={
            "store_id": str(row["store_id"]),
            "advance_id": str(row["id"]),
            "biz_date": _REF_ISO,
            "phone": "13900000000",
            "rebate": "100",
        },
        follow_redirects=True,
    ).get_data(as_text=True)
    assert "芝麻服务费是官方导入的，不能改" in edit
    # 店员不能删
    deleted = c.post(
        "/advance/delete",
        data={"store_id": str(row["store_id"]), "advance_id": str(row["id"])},
        follow_redirects=True,
    ).get_data(as_text=True)
    assert "店员不能删" in deleted
    # 管理员可以删
    c.post("/logout")
    c.post("/login", data={"username": "admin", "pin": "123456"})
    admin_del = c.post(
        "/advance/delete",
        data={"store_id": str(row["store_id"]), "advance_id": str(row["id"])},
        follow_redirects=True,
    ).get_data(as_text=True)
    assert "已删除" in admin_del


def test_sesame_week_bulletin_from_imported_rows(tmp_db, admin_client):
    c = admin_client
    data = _make_sesame_xlsx(_sample_rows())
    c.post("/advance/sesame/preview", data={"sesame_file": (BytesIO(data), "s.xlsx")})
    c.post("/advance/sesame/confirm", follow_redirects=True)
    page = c.get(f"/advance/sesame/week?start={_REF_ISO}&end={_REF_ISO}").get_data(as_text=True)
    assert "芝麻周报" in page
    assert "【芝麻直降办理周报】" in page
    # 通报表门店用全称
    assert "示例市甲街vivo体验店" in page
    assert "9.58" in page
    # 戊店只有退款：净笔数为 -1
    assert "邻市戊路vivo体验店" in page
    city = c.get(f"/advance/sesame/week?start={_REF_ISO}&end={_REF_ISO}&city=示例市").get_data(as_text=True)
    assert "示例市甲街vivo体验店" in city
    assert "邻市戊路vivo体验店" not in city
    xlsx = c.get(f"/advance/sesame/week.xlsx?start={_REF_ISO}&end={_REF_ISO}")
    assert xlsx.status_code == 200
    book = openpyxl.load_workbook(BytesIO(xlsx.get_data()))
    assert book.active["A1"].value == "门店"
    c.post("/logout")
    c.post("/login", data={"username": "alpha", "pin": "123456"})
    blocked = c.get("/advance/sesame/week", follow_redirects=True)
    assert "需要管理员或只读权限" in blocked.get_data(as_text=True)


def _make_orders_xlsx(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["订单号", "冻结金额", "期数", "订单标题"])
    for r in rows:
        ws.append(r)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_tier_category_mapping(tmp_db):
    from app.sesame import TIER_DEFAULT_RULES, tier_category

    rules = TIER_DEFAULT_RULES
    assert tier_category("购机优惠500档（低消70-24期）", rules) == "小天才直降"
    assert tier_category("购机优惠1000档（35元A套-36期）", rules) == "AI手机"
    assert tier_category("购机优惠1400档（低消79元-36期）", rules) == "新用户芝麻直降"
    assert tier_category("新-全品类优惠活动1900档", rules) == "新用户芝麻直降"
    assert tier_category("没有档位的标题", rules) == "未分类"
    # 阈值可配：把小天才档改成 600
    assert tier_category("购机优惠600档", {"xtc": 600, "ai": 1000}) == "小天才直降"


def test_sesame_import_with_orders_and_tier_stats(tmp_db, admin_client):
    """明细 + 订单信息一起导入：通报表按档位列办理笔数，退款列不再出现。"""
    c = admin_client
    detail = _make_sesame_xlsx(_sample_rows())
    orders = _make_orders_xlsx([
        ["1786774661473466851", 480.0, 24, "购机优惠500档（低消70-24期）"],
        ["1785792754901613702", 792.0, 36, "购机优惠1400档（低消79元-36期）"],
        ["1786527145403143969", 1368.0, 36, "购机优惠1900档（低消109元-36期）"],
    ])
    c.post(
        "/advance/sesame/preview",
        data={"sesame_file": (BytesIO(detail), "s.xlsx"), "order_file": (BytesIO(orders), "o.xlsx")},
        content_type="multipart/form-data",
    )
    page = c.get("/advance/sesame").get_data(as_text=True)
    assert "档位覆盖" in page and "小天才直降" in page
    c.post("/advance/sesame/confirm", follow_redirects=True)

    week = c.get(f"/advance/sesame/week?start={_REF_ISO}&end={_REF_ISO}").get_data(as_text=True)
    # 通报表列：小天才 / AI手机 / 直降分档；退款列不再出现
    assert "芝麻直降办理周报" in week
    assert ">小天才</th>" in week and ">AI手机</th>" in week
    assert ">1400档</th>" in week and ">1900档</th>" in week
    assert "退款</th>" not in week
    # 甲店小天才 1 笔；净笔数=扣费−退款（戊店 0−1=−1）
    assert "净办理笔数" in week and "服务费净额" in week
    textarea = week.split('id="reviewText"')[1].split("</textarea>")[0]
    assert "【芝麻直降办理周报】" in textarea
    assert "净办理" in textarea and "退款" not in textarea
    # Excel 按档位 sheet：直降分档办理笔数（1900 档退款不算办理）
    xlsx = c.get(f"/advance/sesame/week.xlsx?start={_REF_ISO}&end={_REF_ISO}")
    book = openpyxl.load_workbook(BytesIO(xlsx.get_data()))
    assert "按档位" in book.sheetnames
    ws = book["按档位"]
    tiers = {ws.cell(r, 1).value: (ws.cell(r, 3).value, ws.cell(r, 4).value) for r in range(2, ws.max_row + 1)}
    assert tiers["1400档"] == (1, 5.54)
    # 1900 档本期只有退款（原扣费在上期）：净办理 -1 笔，净 -9.58
    assert tiers["1900档"] == (-1, -9.58)


def test_sesame_month_mode_and_filters(tmp_db, admin_client):
    """月报模式：标题、区域经理筛选；不再出地市/区域经理独立表。"""
    c = admin_client
    data = _make_sesame_xlsx(_sample_rows())
    c.post("/advance/sesame/preview", data={"sesame_file": (BytesIO(data), "s.xlsx")})
    c.post("/advance/sesame/confirm", follow_redirects=True)
    page = c.get(f"/advance/sesame/week?mode=month&start={_MONTH_START}").get_data(as_text=True)
    assert "芝麻直降办理月报" in page
    assert "区域经理" in page  # 筛选器
    assert "按地市" not in page  # 独立分类表已撤，改筛选
    assert "示例市甲街vivo体验店" in page
    # 地市筛选仍可用
    city = c.get(f"/advance/sesame/week?mode=month&start={_MONTH_START}&city=示例市").get_data(as_text=True)
    assert "示例市甲街vivo体验店" in city
    assert "示例市丁路vivo体验店" not in city
    xlsx = c.get(f"/advance/sesame/week.xlsx?mode=month&start={_MONTH_START}")
    assert xlsx.status_code == 200
    book = openpyxl.load_workbook(BytesIO(xlsx.get_data()))
    assert "按地市" not in book.sheetnames
    assert "按档位" in book.sheetnames


def test_sesame_shows_in_advance_totals(tmp_db, admin_client):
    c = admin_client
    data = _make_sesame_xlsx(_sample_rows()[:1])
    c.post("/advance/sesame/preview", data={"sesame_file": (BytesIO(data), "s.xlsx")})
    c.post("/advance/sesame/confirm", follow_redirects=True)
    page = c.get("/advance").get_data(as_text=True)
    assert "芝麻服务费" in page
    # 示例甲店 9.58
    with db.get_db() as conn:
        totals = db.advance_month_totals(
            conn,
            [conn.execute("SELECT id FROM stores WHERE code='store-alpha'").fetchone()["id"]],
            db.today_local(),
        )
    assert abs(totals[list(totals)[0]]["sesame"] - 9.58) < 0.01


def test_sesame_export_has_column(tmp_db, admin_client):
    c = admin_client
    data = _make_sesame_xlsx(_sample_rows()[:1])
    c.post("/advance/sesame/preview", data={"sesame_file": (BytesIO(data), "s.xlsx")})
    c.post("/advance/sesame/confirm", follow_redirects=True)
    resp = c.get(f"/advance.xlsx?month={db.today_local().strftime('%Y-%m')}")
    wb = openpyxl.load_workbook(BytesIO(resp.get_data()))
    summary = wb["汇总表"]
    headers = [summary.cell(2, col).value for col in range(1, 7)]
    assert "芝麻服务费" in headers
    # 门店明细表也有
    store_sheet = wb["示例甲店"]
    store_headers = [store_sheet.cell(2, col).value for col in range(1, 11)]
    assert "芝麻服务费" in store_headers


def test_import_more_than_200_rows_imports_all(tmp_db, admin_client):
    """回归：预览只展示前 200 笔，确认导入必须全量；全店垫资页也不能截断。"""
    c = admin_client
    rows = []
    for i in range(230):
        rows.append(
            [
                f"TESTEXT{i:08d}", f"TESTORD{i:08d}", "江苏省", "示例市", "甲区",
                "JSCM_20250116110103117937984", "91320602MA7E6JC70A", "示例公司甲",
                "JSCM_10000001", "示例甲店vivo专营店", "加盟店", _REF_MONTH,
                1000.00, -5.00, "处理成功",
                f"行业芝麻订单(TESTEXT{i:08d})服务费", f"{_REF_ISO} 14:19:38.0",
            ]
        )
    # 预览：ready_count 显示全量 230，预览表只展示 200
    resp = c.post(
        "/advance/sesame/preview",
        data={"sesame_file": (BytesIO(_make_sesame_xlsx(rows)), "s.xlsx")},
        follow_redirects=True,
    )
    page = resp.get_data(as_text=True)
    assert "可导入 <strong>230</strong>" in page
    assert "仅展示前 200 笔" in page
    with db.get_db() as conn:
        assert conn.execute("SELECT COUNT(*) FROM advance_posts WHERE source='sesame'").fetchone()["COUNT(*)"] == 0
    # 确认：全部 230 笔都要进库
    c.post("/advance/sesame/confirm", follow_redirects=True)
    with db.get_db() as conn:
        imported = conn.execute(
            "SELECT COUNT(*), SUM(sesame) FROM advance_posts WHERE source='sesame'"
        ).fetchone()
    assert imported["COUNT(*)"] == 230
    assert imported["SUM(sesame)"] == 230 * 500  # 服务费 -5.00 元→入账 +5.00，存 +500 分/笔
    # 全店视图：应显示全部 230 笔，而不是被截断到 200
    page = c.get("/advance?store_id=all").get_data(as_text=True)
    assert "230 笔" in page  # 列表计数用全量
    # 每行备注（芝麻服务费 + 订单号）渲染约 2 处；230 行远超市旧 200 条上限，
    # 故多过 200*2=400 即证明未被截断。
    assert page.count("TESTORD") > 400


def test_sesame_advance_cannot_unpay(tmp_db, admin_client):
    """芝麻导入的垫资「导入即已兑」，取消兑付对它无效。"""
    from datetime import date

    with db.get_db() as conn:
        sid = conn.execute("SELECT id FROM stores WHERE code='store-alpha'").fetchone()["id"]
        admin_id = conn.execute("SELECT id FROM users WHERE username='admin'").fetchone()["id"]
        aid = db.record_advance(
            conn, store_id=sid, user_id=admin_id, biz_date=date.today(),
            sesame=5.5, source="sesame", ext_id="TEST_R1", paid=True,
        )
        db.set_advance_paid(conn, [aid], paid=False, user_id=admin_id)
        row = conn.execute("SELECT paid FROM advance_posts WHERE id=?", (aid,)).fetchone()
        assert int(row["paid"] or 0) == 1

