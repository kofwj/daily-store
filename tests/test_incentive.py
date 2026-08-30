from datetime import date

from app.incentive import DEFAULTS, judge, judge_with_advisor, judge_without_advisor, money_text
from app.metrics_seed import rollup_amount
from app.settlement import prev_month_start


def test_with_advisor_table():
    """未达标按接近度分两档：总量差 ≤3 算差一点，罚得轻；差得远顶格。"""
    assert judge_with_advisor(3, 7)["store_reward"] == 500
    assert judge_with_advisor(2, 8)["store_reward"] == 200
    row = judge_with_advisor(0, 10)
    assert row["store_reward"] == 0
    assert row["advisor_penalty"] == 100
    # AI≥5：顾问免责，罚门店；差一点 100，差得远 200
    row = judge_with_advisor(5, 2)  # 总量 7，差 3 → 差一点
    assert row["store_penalty"] == 100
    assert row["advisor_penalty"] == 0
    row = judge_with_advisor(5, 0)  # 总量 5，差 5 → 差得远
    assert row["store_penalty"] == 200
    assert row["advisor_penalty"] == 0
    # AI 1-4：整体欠佳；差一点 100/50，差得远 200/50
    row = judge_with_advisor(2, 7)  # 总量 9，差 1
    assert row["store_penalty"] == 100
    assert row["advisor_penalty"] == 50
    row = judge_with_advisor(2, 3)  # 总量 5，差 5
    assert row["store_penalty"] == 200
    assert row["advisor_penalty"] == 50
    # AI=0：整体极差；差一点顾问罚减半，差得远顶格
    row = judge_with_advisor(0, 8)  # 总量 8，差 2
    assert row["store_penalty"] == 200
    assert row["advisor_penalty"] == 50
    row = judge_with_advisor(0, 4)  # 总量 4，差 6
    assert row["store_penalty"] == 200
    assert row["advisor_penalty"] == 100


def test_near_miss_gap_configurable():
    rules = dict(DEFAULTS)
    rules["near_miss_gap"] = 1  # 只有差 1 分算差一点
    row = judge_with_advisor(5, 2, rules)  # 总量 7，差 3 → 差得远
    assert row["store_penalty"] == 200
    rules["near_miss_gap"] = 5  # 差 5 以内都算差一点
    row = judge_with_advisor(2, 4, rules)  # 总量 6，差 4 → 差一点
    assert row["store_penalty"] == 100
    assert row["advisor_penalty"] == 50


def test_near_miss_effective_from_september():
    """生效月守卫：9 月前的考核月走旧阶梯，不回头改当月数字。"""
    from datetime import date

    from app.incentive import NEAR_MISS_FROM

    before = date(2026, 8, 31)
    # 8 月：AI5+直降2 仍按旧档罚门店 200
    row = judge(True, 5, 2, as_of=before)
    assert row["store_penalty"] == 200
    row = judge(True, 2, 3, as_of=before)
    assert row["store_penalty"] == 100
    row = judge(True, 0, 8, as_of=before)
    assert row["advisor_penalty"] == 100
    # 9 月起走新阶梯
    row = judge(True, 5, 2, as_of=NEAR_MISS_FROM)
    assert row["store_penalty"] == 100
    row = judge(True, 2, 3, as_of=NEAR_MISS_FROM)
    assert row["store_penalty"] == 200
    row = judge(True, 0, 8, as_of=NEAR_MISS_FROM)
    assert row["advisor_penalty"] == 50


def test_without_advisor_table():
    assert judge_without_advisor(1, 1)["store_reward"] == 200
    assert judge_without_advisor(2, 0)["store_penalty"] == 50
    assert judge_without_advisor(0, 3)["store_penalty"] == 50
    assert judge_without_advisor(0, 0)["store_penalty"] == 100


def test_custom_rules_change_threshold():
    rules = dict(DEFAULTS)
    rules["total_threshold"] = 12
    rules["reward_best"] = 800
    assert judge(True, 3, 7, rules)["passed"] is False
    assert judge(True, 5, 7, rules)["store_reward"] == 800


def test_new_user_cut_includes_full_category():
    month_vals = {
        "coin_cut_new_recharge": 0,
        "coin_cut_new_sesame": 0,
        "coin_cut_new_savings": 0,
        "coin_cut_new_full": 1,
        "coin_cut_old": 4,
        "coin_cut_xtc": 2,
    }
    new_cut = rollup_amount(month_vals, "coin_cut")
    assert new_cut == 1
    row = judge(False, 1, new_cut)
    assert row["new_cut"] == 1
    assert row["passed"] is True
    assert row["goal"] == "AI、新用户直降均破 0"
    assert judge(True, 0, 10)["label"] == "总量靠直降"


def test_money_text():
    assert money_text(judge(True, 3, 7)) == "奖门店 500"
    assert money_text(judge(False, 0, 0)) == "罚门店 100"
    assert "罚门店 100" in money_text(judge(True, 2, 7))  # 差一点的欠佳档
    assert "罚顾问 50" in money_text(judge(True, 2, 7))


def test_settlement_formulas_follow_confirmed_rules():
    from app.settlement import bonus_formula, commission_target, focus_score_formula

    assert commission_target("A") == 4000
    assert commission_target("B") == 1000
    assert "I5" in bonus_formula("I5", "A") and "4000" in bonus_formula("I5", "A")
    assert "1000" in bonus_formula("I6", "B") and "1200" in bonus_formula("I6", "B")
    assert "M5>=10" in focus_score_formula("M5", "A")
    assert "M6>=4" in focus_score_formula("M6", "B")


def test_settlement_invoice_uses_previous_month():
    assert prev_month_start(date(2026, 8, 18)) == date(2026, 7, 1)
    assert prev_month_start(date(2026, 1, 5)) == date(2025, 12, 1)


def test_incentive_xlsx_exports_draft(tmp_db):
    from io import BytesIO

    import openpyxl

    from app.web import create_app

    app = create_app()
    app.config["TESTING"] = True
    c = app.test_client()
    c.post("/login", data={"username": "admin", "pin": "123456"})
    r = c.get("/incentive.xlsx")
    assert r.status_code == 200
    assert "filename=settlement_" in r.headers.get("Content-Disposition", "")
    wb = openpyxl.load_workbook(BytesIO(r.get_data()))
    ws = wb["移动接入"]
    assert "运营商绩效" in (ws["A1"].value or "")
    assert ws["E2"].value == "酬金（50%）"
    assert ws["S2"].value == "考核奖惩"
    # 第一家店：实际酬金公式统一为 开票+房补-垫资
    assert ws["I5"].value == "=F5+G5-H5"
    # 导出门店用官方全称，不用短名
    assert ws["B5"].value == "示例市甲街vivo体验店"
    # 黄色手填格
    assert ws["F5"].fill.fgColor.rgb[-6:] == "FFF2CC"
    # 口径页
    assert "口径" in wb.sheetnames


def test_unreported_store_is_not_penalized(tmp_db):
    """本月没交过日报的店不参与考核，奖惩记 0。"""
    from datetime import date

    from app import db as _db
    from app.helpers import store_forecast
    from app.web import create_app

    app = create_app()
    app.config["TESTING"] = True
    c = app.test_client()
    c.post("/login", data={"username": "admin", "pin": "123456"})
    with _db.get_db() as conn:
        store = conn.execute("SELECT * FROM stores WHERE code='store-epsilon'").fetchone()
        judged = store_forecast(conn, store, date.today())
        assert judged["reported"] is False
        assert judged["label"] == "本月未交"
        assert judged["net"] == 0
        assert judged["store_penalty"] == 0
    page = c.get("/incentive").get_data(as_text=True)
    assert "本月未交" in page
    assert "邻市戊路vivo体验店" in page
